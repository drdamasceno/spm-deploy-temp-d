#!/usr/bin/env python3
"""Script para inspecionar estrutura dos arquivos reais."""

import sys
import openpyxl
from pathlib import Path

PP_PATH = "/mnt/user-data/uploads/SPM_-_FB_-_Conf__01_a_09_04_2026_Claudexlsx.xlsx"
OFX_PATH = "/mnt/user-data/uploads/EXTRATO_POR_PERIODO_120426_221307.ofx"

def inspect_pp():
    print("=== INSPECIONANDO PP XLSX ===")
    wb = openpyxl.load_workbook(PP_PATH, data_only=True)
    ws = wb.active
    print(f"Planilha ativa: {ws.title}")
    print(f"Dimensões: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()
    print("Primeiras 30 linhas:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        non_none = [c for c in row if c is not None]
        if non_none:
            print(f"  Linha {i}: {row[:10]}")
    
    # Find all "Total" rows
    print()
    print("Linhas com 'Total':")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] and str(row[0]).strip().lower() == 'total':
            print(f"  Linha {i}: {row[:10]}")
        if i > 200:
            break

def inspect_ofx():
    print()
    print("=== INSPECIONANDO OFX ===")
    with open(OFX_PATH, 'rb') as f:
        raw = f.read()
    # Try to detect encoding
    for enc in ['latin-1', 'cp1252', 'utf-8']:
        try:
            text = raw.decode(enc)
            print(f"Encoding: {enc}")
            print(f"Tamanho: {len(text)} chars")
            # Show first 80 lines
            lines = text.split('\n')[:80]
            for j, line in enumerate(lines, 1):
                print(f"  L{j}: {line.rstrip()}")
            break
        except:
            continue

if __name__ == '__main__':
    inspect_pp()
    inspect_ofx()
