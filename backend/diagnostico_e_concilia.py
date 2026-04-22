#!/usr/bin/env python3
"""
Diagnostico e execucao da conciliacao SPM.
Execute do diretorio ~/spm-sistemafinanceiro/:
    python backend/diagnostico_e_concilia.py
"""
import sys, subprocess, json, re
from pathlib import Path

PP_PATH = "/mnt/user-data/uploads/SPM_-_FB_-_Conf__01_a_09_04_2026_Claudexlsx.xlsx"
OFX_PATH = "/mnt/user-data/uploads/EXTRATO_POR_PERIODO_120426_221307.ofx"

def install_deps():
    for pkg in ["openpyxl"]:
        try:
            __import__(pkg)
        except ImportError:
            print("Instalando {}...".format(pkg))
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

def inspect_ofx(path):
    """Mostra as primeiras 80 linhas do OFX."""
    print("\n=== ESTRUTURA OFX ===")
    try:
        raw = Path(path).read_bytes()
        for enc in ["latin-1", "cp1252", "utf-8"]:
            try:
                text = raw.decode(enc)
                print("Encoding: {}, tamanho: {} chars".format(enc, len(text)))
                lines = text.split("\n")[:80]
                for i, line in enumerate(lines, 1):
                    print("  L{:3}: {}".format(i, line.rstrip()))
                # Count transactions
                n_trans = len(re.findall(r"<STMTTRN>", text, re.IGNORECASE))
                print("\nTotal transacoes STMTTRN: {}".format(n_trans))
                break
            except UnicodeDecodeError:
                continue
    except Exception as e:
        print("Erro ao ler OFX: {}".format(e))

def inspect_pp(path):
    """Mostra as primeiras 30 linhas do PP XLSX."""
    print("\n=== ESTRUTURA PP XLSX ===")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        print("Planilha: {}, Linhas: {}, Colunas: {}".format(ws.title, ws.max_row, ws.max_column))
        print("\nPrimeiras 30 linhas:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            non_none = [v for v in row if v is not None]
            if non_none:
                print("  Linha {:3}: {}".format(i, str(row[:10])[:120]))
        # Count "Total" rows
        totais = 0
        nomes = []
        for row in ws.iter_rows(values_only=True):
            if row[0] and str(row[0]).strip().lower() == "total":
                totais += 1
            # Detectar cabecalhos de prestador (linha A)
            vals_nao_nulos = [v for v in row if v is not None]
            if len(vals_nao_nulos) == 1 and " - " in str(vals_nao_nulos[0])[:100]:
                nomes.append(str(vals_nao_nulos[0])[:80])
        print("\nLinhas Total encontradas: {}".format(totais))
        print("Prestadores detectados (primeiros 10):".format())
        for n in nomes[:10]:
            print("  ", n)
    except Exception as e:
        print("Erro ao ler XLSX: {}".format(e))
        import traceback
        traceback.print_exc()

def run_conciliacao():
    """Executa a conciliacao com os arquivos reais."""
    print("\n=== EXECUTANDO CONCILIACAO ===")
    sys.path.insert(0, str(Path(__file__).parent / "src"))
    try:
        from pega_plantao import parse_relatorio
        from extrato_bradesco import parse_extrato
        from conciliacao_spm import conciliar
    except ImportError as e:
        print("Erro de import: {}. Certifique-se de estar em ~/spm-sistemafinanceiro/".format(e))
        return
    
    print("Parseando PP...")
    try:
        pp = parse_relatorio(PP_PATH)
        print("PP: {} registros".format(len(pp)))
        for r in pp[:3]:
            print("  ", r["nome_prestador"], "|", r["contrato"], "| saldo=", r["saldo"])
    except Exception as e:
        print("ERRO PP: {}".format(e))
        import traceback; traceback.print_exc()
        return
    
    print("\nParseando Extrato OFX...")
    try:
        extrato = parse_extrato(OFX_PATH)
        print("Extrato: {} transacoes".format(len(extrato)))
        for t in extrato[:3]:
            print("  ", t["data"], "|", t["tipo"], "|", t["valor"], "|", t["memo"][:40])
    except Exception as e:
        print("ERRO Extrato: {}".format(e))
        import traceback; traceback.print_exc()
        return
    
    print("\nConciliando...")
    resultado = conciliar(pp, extrato)
    resumo = resultado["resumo"]
    
    print("\n=== RESUMO ===")
    for k, v in resumo.items():
        print("  {}: {}".format(k, v))
    
    # Gerar relatorio
    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(exist_ok=True)
    
    sys.path.insert(0, str(Path(__file__).parent))
    try:
        from conciliar_spm import gerar_relatorio_xlsx
        xlsx = gerar_relatorio_xlsx(resultado, PP_PATH, OFX_PATH, output_dir)
        if xlsx:
            print("\nRelatorio salvo: {}".format(xlsx))
    except Exception as e:
        print("Erro ao gerar XLSX: {}".format(e))

if __name__ == '__main__':
    install_deps()
    inspect_ofx(OFX_PATH)
    inspect_pp(PP_PATH)
    run_conciliacao()
