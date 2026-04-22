#!/usr/bin/env python3
"""
CLI Conciliacao SPM.

Uso:
    python conciliar_spm.py <caminho_pp.xlsx> <caminho_extrato.ofx>

Fluxo:
    1. Parsear PP XLSX -> pp_data
    2. Parsear Extrato OFX -> extrato_data
    3. Conciliar -> resultado
    4. Imprimir resumo no terminal
    5. Gerar relatorio XLSX em backend/output/conciliacao_MMAAAA.xlsx

Relatorio XLSX:
    - Aba "Conciliacao": todos os registros com status colorido
      Verde: MATCH_AUTOMATICO, FRACIONADO, CONCILIADO_CATEGORIA
      Amarelo: MANUAL_PENDENTE
      Vermelho: NAO_CLASSIFICADO
    - Aba "Pendencias Hugo": MANUAL_PENDENTE + NAO_CLASSIFICADO
    - Aba "Log": data/hora, arquivos, totais
"""

import sys
import json
import logging
import argparse
from pathlib import Path
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger('conciliar_spm')

sys.path.insert(0, str(Path(__file__).parent / "src"))

from pega_plantao import parse_relatorio
from extrato_bradesco import parse_extrato
from conciliacao_spm import conciliar


def _formato_moeda(valor: float) -> str:
    return "R$ {:,.2f}".format(valor).replace(",", "X").replace(".", ",").replace("X", ".")


def gerar_relatorio_xlsx(resultado: dict, pp_path: str, ofx_path: str, output_dir: Path) -> Path:
    """Gera relatorio XLSX com 3 abas."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl nao instalado. Execute: pip install openpyxl")
        return None

    resumo = resultado["resumo"]
    registros = resultado["registros"]
    extras = resultado.get("transacoes_extras", [])

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    laranja = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    cores = {
        "MATCH_AUTOMATICO": verde,
        "FRACIONADO": verde,
        "CONCILIADO_CATEGORIA": verde,
        "MANUAL_PENDENTE": amarelo,
        "NAO_CLASSIFICADO": vermelho,
        "SEM_MOVIMENTO": cinza,
        "SALDO_NEGATIVO": laranja,
    }

    meses = [r["mes_competencia"] for r in registros if r.get("mes_competencia")]
    periodo = meses[0] if meses else datetime.now().strftime("%m/%Y")
    periodo_nome = periodo.replace("/", "")

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "conciliacao_{}.xlsx".format(periodo_nome)

    wb = openpyxl.Workbook()

    # === ABA 1: Conciliacao ===
    ws1 = wb.active
    ws1.title = "Conciliacao"
    headers = [
        "Nome Prestador", "Contrato", "Mes Competencia",
        "Saldo PP (R$)", "Status", "Categoria",
        "PIX Matched (qtd)", "Valor PIX Total (R$)", "Divergencia (R$)",
        "Tipo Doc", "Documento", "Chave PIX"
    ]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    for reg in registros:
        pix_qtd = len(reg.get("pix_matched", []))
        row = [
            reg["nome_prestador"], reg["contrato"], reg["mes_competencia"],
            reg["saldo_pp"], reg["status"], reg.get("categoria", ""),
            pix_qtd, reg.get("valor_pix_total", 0.0), reg.get("divergencia", 0.0),
            reg.get("tipo_doc", ""), reg.get("documento", ""), reg.get("chave_pix", ""),
        ]
        ws1.append(row)
        fill = cores.get(reg["status"])
        if fill:
            for cell in ws1[ws1.max_row]:
                cell.fill = fill

    for i in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 22

    # === ABA 2: Pendencias Hugo ===
    ws2 = wb.create_sheet("Pendencias Hugo")
    pendentes = [r for r in registros if r["status"] in ("MANUAL_PENDENTE", "NAO_CLASSIFICADO")]
    ws2.append(["ATENCAO: Requer aprovacao de Hugo antes de qualquer pagamento"])
    ws2.append([])
    headers2 = ["Nome Prestador", "Contrato", "Mes", "Saldo PP (R$)", "Status",
                "PIX Mais Proximo", "Valor PIX (R$)", "Divergencia (R$)", "Acao"]
    ws2.append(headers2)
    for cell in ws2[3]:
        cell.font = Font(bold=True)
    for reg in pendentes:
        pix_proximo, valor_pix = "", 0.0
        if reg.get("pix_matched"):
            p = reg["pix_matched"][0]
            pix_proximo = "{} | {} | {}".format(
                p.get("data", ""), p.get("titular_pix", ""), p.get("memo", "")[:30]
            )
            valor_pix = abs(p.get("valor", 0.0))
        acao = "Verificar titular e valor" if reg["status"] == "MANUAL_PENDENTE" else "Sem match"
        ws2.append([
            reg["nome_prestador"], reg["contrato"], reg["mes_competencia"],
            reg["saldo_pp"], reg["status"], pix_proximo, valor_pix,
            reg.get("divergencia", 0.0), acao
        ])
        fill = cores.get(reg["status"])
        if fill:
            for cell in ws2[ws2.max_row]:
                cell.fill = fill
    for i in range(1, 10):
        ws2.column_dimensions[get_column_letter(i)].width = 25

    # === ABA 3: Anomalias (SALDO_NEGATIVO) ===
    ws_anom = wb.create_sheet("Anomalias")
    anomalias = [r for r in registros if r["status"] == "SALDO_NEGATIVO"]
    ws_anom.append(["ANOMALIAS DO PP — saldo negativo (taxa SPM sem plantao positivo, estorno, etc.)"])
    ws_anom.append(["NAO sao pendencia de pagamento — auditar com Pega Plantao"])
    ws_anom.append([])
    headers_anom = ["Nome Prestador", "Contrato", "Mes", "Saldo PP (R$)", "Tipo Doc", "Documento"]
    ws_anom.append(headers_anom)
    for cell in ws_anom[4]:
        cell.font = Font(bold=True)
    for reg in anomalias:
        ws_anom.append([
            reg["nome_prestador"], reg["contrato"], reg["mes_competencia"],
            reg["saldo_pp"], reg.get("tipo_doc", ""), reg.get("documento", ""),
        ])
        fill = cores.get(reg["status"])
        if fill:
            for cell in ws_anom[ws_anom.max_row]:
                cell.fill = fill
    for i in range(1, len(headers_anom) + 1):
        ws_anom.column_dimensions[get_column_letter(i)].width = 25

    # === ABA 4: Sem Movimento (auditoria) ===
    ws_sm = wb.create_sheet("Sem Movimento")
    sem_mov = [r for r in registros if r["status"] == "SEM_MOVIMENTO"]
    ws_sm.append(["PRESTADORES SEM MOVIMENTO — saldo zero (sem plantao no periodo)"])
    ws_sm.append(["Lista de auditoria, nao requer acao"])
    ws_sm.append([])
    headers_sm = ["Nome Prestador", "Contrato", "Mes", "Tipo Doc", "Documento"]
    ws_sm.append(headers_sm)
    for cell in ws_sm[4]:
        cell.font = Font(bold=True)
    for reg in sem_mov:
        ws_sm.append([
            reg["nome_prestador"], reg["contrato"], reg["mes_competencia"],
            reg.get("tipo_doc", ""), reg.get("documento", ""),
        ])
        fill = cores.get(reg["status"])
        if fill:
            for cell in ws_sm[ws_sm.max_row]:
                cell.fill = fill
    for i in range(1, len(headers_sm) + 1):
        ws_sm.column_dimensions[get_column_letter(i)].width = 25

    # === ABA 5: Log ===
    ws3 = wb.create_sheet("Log")
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws3.append(["SPM Sistema Financeiro - Log de Conciliacao"])
    ws3.append([])
    ws3.append(["Data/Hora Execucao:", agora])
    ws3.append(["Arquivo PP:", str(pp_path)])
    ws3.append(["Arquivo Extrato:", str(ofx_path)])
    ws3.append([])
    ws3.append(["=== RESUMO ==="])
    for k, v in resumo.items():
        label = k.replace("_", " ").title()
        if "valor" in k.lower():
            ws3.append([label, _formato_moeda(float(v))])
        elif "percentual" in k.lower():
            ws3.append([label, "{:.1f}%".format(v)])
        else:
            ws3.append([label, v])
    ws3.append([])
    ws3.append(["=== TRANSACOES EXTRAS ==="])
    ws3.append(["Data", "Memo", "Valor", "Tipo", "Categoria"])
    for t in extras:
        ws3.append([t.get("data"), t.get("memo", "")[:50], t.get("valor"), t.get("tipo"), t.get("categoria")])
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 55

    wb.save(output_path)
    logger.info("Relatorio salvo: {}".format(output_path))
    return output_path


def imprimir_resumo(resultado: dict, registros: list):
    """Imprime resumo formatado no terminal."""
    r = resultado["resumo"]
    print()
    print("=" * 60)
    print("  SPM SISTEMA FINANCEIRO - RESULTADO DA CONCILIACAO")
    print("=" * 60)
    print("  Registros PP:        {:>5}".format(r["total_registros_pp"]))
    print("  PIX no extrato:      {:>5}".format(r["total_pix_extrato"]))
    print()
    print("  OK Match Automatico: {:>5}".format(r["match_automatico"]))
    print("  OK Fracionado:       {:>5}".format(r["fracionado"]))
    print("  OK Concil. Categ.:   {:>5}".format(r["conciliado_categoria"]))
    print("  ?? Manual Pendente:  {:>5}".format(r["manual_pendente"]))
    print("  XX Nao Classificado: {:>5}".format(r["nao_classificado"]))
    print("  -- Sem Movimento:    {:>5}  (saldo=0, excluido do %)".format(r.get("sem_movimento", 0)))
    print("  !! Saldo Negativo:   {:>5}  (anomalia PP, excluido do %)".format(r.get("saldo_negativo", 0)))
    print()
    print("  Valor Total PP:      R$ {:>12,.2f}".format(r["valor_total_pp"]))
    print("  Valor Conciliado:    R$ {:>12,.2f}".format(r["valor_conciliado"]))
    print("  Percentual:          {:>7.1f}%".format(r["percentual_conciliado"]))
    print("=" * 60)
    pendentes = [reg for reg in registros if reg["status"] in ("MANUAL_PENDENTE", "NAO_CLASSIFICADO")]
    if pendentes:
        print()
        print("PENDENCIAS PARA HUGO:")
        print("-" * 60)
        for reg in pendentes:
            pix_info = ""
            if reg.get("pix_matched"):
                p = reg["pix_matched"][0]
                pix_info = " <- PIX {} {} R${:,.2f}".format(
                    p.get("data", ""), p.get("titular_pix", "")[:20], abs(p.get("valor", 0)))
            print("  [{}] {:<30} R$ {:>9,.2f}{}".format(
                reg["status"][:4], reg["nome_prestador"][:30],
                reg["saldo_pp"], pix_info))
    print()


def main():
    parser = argparse.ArgumentParser(description="Conciliacao SPM")
    parser.add_argument("pp", help="Caminho para o arquivo PP XLSX")
    parser.add_argument("extrato", help="Caminho para o arquivo Extrato OFX")
    parser.add_argument("--excecoes", help="JSON com excecoes PJ", default=None)
    parser.add_argument("--output", help="Diretorio de saida", default="backend/output")
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)

    excecoes_pj = {}
    if args.excecoes:
        try:
            with open(args.excecoes) as f:
                excecoes_pj = json.load(f)
            logger.info("Excecoes PJ: {}".format(list(excecoes_pj.keys())))
        except Exception as e:
            logger.warning("Erro ao carregar excecoes: {}".format(e))

    logger.info("Parseando PP: {}".format(args.pp))
    try:
        pp_data = parse_relatorio(args.pp)
    except Exception as e:
        logger.error("Erro ao parsear PP: {}".format(e))
        sys.exit(1)
    logger.info("PP: {} registros".format(len(pp_data)))

    logger.info("Parseando Extrato: {}".format(args.extrato))
    try:
        extrato_data = parse_extrato(args.extrato)
    except Exception as e:
        logger.error("Erro ao parsear extrato: {}".format(e))
        sys.exit(1)
    logger.info("Extrato: {} transacoes".format(len(extrato_data)))

    logger.info("Executando conciliacao...")
    resultado = conciliar(pp_data, extrato_data, excecoes_pj)

    imprimir_resumo(resultado, resultado["registros"])

    output_dir = Path(args.output)
    xlsx_path = gerar_relatorio_xlsx(resultado, args.pp, args.extrato, output_dir)
    if xlsx_path:
        print("Relatorio salvo em: {}".format(xlsx_path))

    pendencias = resultado["resumo"]["manual_pendente"] + resultado["resumo"]["nao_classificado"]
    if pendencias > 0:
        print("ATENCAO: {} itens aguardam aprovacao de Hugo.".format(pendencias))
    sys.exit(0)


if __name__ == '__main__':
    main()
