# spm-faturas: Gerador de planilhas Excel

from __future__ import annotations

import logging
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

CORES = {
    "header":     "1A2F5A",
    "verde_l":    "E2EFDA",
    "vermelho_l": "FCE4D6",
    "amarelo":    "FFF2CC",
    "azul_l":     "D6E4F0",
    "verde_pj":   "2E7D32",
    "azul_pj":    "1F6AA5",
    "cinza":      "F2F2F2",
    "previa":     "FFF3CD",
}

FILL_HEADER  = PatternFill("solid", fgColor=CORES["header"])
FILL_VERDE   = PatternFill("solid", fgColor=CORES["verde_l"])
FILL_VERMELHO = PatternFill("solid", fgColor=CORES["vermelho_l"])
FILL_AMARELO = PatternFill("solid", fgColor=CORES["amarelo"])
FILL_AZUL    = PatternFill("solid", fgColor=CORES["azul_l"])
FILL_CINZA   = PatternFill("solid", fgColor=CORES["cinza"])
FILL_PREVIA  = PatternFill("solid", fgColor=CORES["previa"])

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD   = Font(bold=True, size=11)
FONT_NORMAL = Font(size=11)
FONT_SUBTOTAL_PJ = Font(bold=True, color="FFFFFF", size=11)
FONT_SUBTOTAL_PF = Font(bold=True, color="FFFFFF", size=11)

FILL_SUBTOTAL_PJ = PatternFill("solid", fgColor=CORES["verde_pj"])
FILL_SUBTOTAL_PF = PatternFill("solid", fgColor=CORES["azul_pj"])

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")

FMT_BRL = '#,##0.00'

# Tipos que contam para PF/PJ (exclui apenas pagamento)
TIPOS_PF_PJ = ("vista", "parcelada", "encargo")


# ---------------------------------------------------------------------------
# API publica
# ---------------------------------------------------------------------------

def gerar_excel(
    fatura: dict,
    classificacao: dict,
    config: dict,
    caminho_saida: str,
) -> str:
    """Entry point. Gera o arquivo Excel completo e retorna o caminho final."""
    wb = Workbook()

    banco = fatura.get("banco", "BB")
    resumo = fatura.get("resumo", {})
    mes_referencia = resumo.get("mes_referencia", "")

    # Unificar lancamentos: classificados + pendentes (manter ordem original)
    todos = classificacao.get("classificados", []) + classificacao.get("pendentes", [])
    todos.sort(key=lambda l: l.get("data", ""))

    stats = classificacao.get("stats", {})
    pendentes = stats.get("pendentes", 0)

    # Aba 1: Lancamentos
    _aba_lancamentos(wb, todos, banco)

    # Aba 2: Parcelamentos
    parceladas = [l for l in todos if l.get("tipo") == "parcelada"]
    _aba_parcelamentos(wb, parceladas, mes_referencia)

    # Aba 3: Projecao Faturas
    _aba_projecao(wb, parceladas, mes_referencia)

    # Aba 4: Resumo Fatura
    _aba_resumo(wb, resumo, stats, banco)

    # Aba 5: PF vs PJ (so se completo)
    gerou_pf_pj = _aba_pf_pj(wb, todos, apenas_se_completo=True)

    # Aba 6: Relatorio Conselho (so se completo)
    gerou_conselho = _aba_relatorio_conselho(wb, todos, resumo, apenas_se_completo=True)

    if pendentes > 0:
        logger.info(
            "Abas 'PF vs PJ' e 'Relatório Conselho' não geradas — "
            "%d lançamentos pendentes de classificação manual.",
            pendentes,
        )

    # Remover sheet default vazia se existir
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # Salvar
    path = Path(caminho_saida)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info("Excel salvo em %s", caminho_saida)
    return str(path)


# ---------------------------------------------------------------------------
# Abas
# ---------------------------------------------------------------------------

def _aba_lancamentos(wb: Workbook, lancamentos: list[dict], banco: str) -> None:
    """Aba 'Lancamentos [BANCO]' — todos os lancamentos.
    Coluna PF/PJ amarela para preenchimento manual dos pendentes.
    """
    titulo = f"Lancamentos {banco}"
    ws = wb.create_sheet(title=titulo)

    headers = [
        "Data", "Descrição", "Cidade", "País", "Tipo",
        "Valor (R$)", "PF/PJ", "Categoria", "Confiança", "Origem",
        "Parcela", "Total Parcelas",
    ]
    _escrever_header(ws, headers)

    for i, lanc in enumerate(lancamentos, start=2):
        tipo = lanc.get("tipo", "")
        origem = lanc.get("origem", "pendente")
        is_pendente = origem == "pendente"

        ws.cell(row=i, column=1, value=lanc.get("data", ""))
        ws.cell(row=i, column=2, value=lanc.get("descricao", ""))
        ws.cell(row=i, column=3, value=lanc.get("cidade", ""))
        ws.cell(row=i, column=4, value=lanc.get("pais", ""))
        ws.cell(row=i, column=5, value=tipo)
        ws.cell(row=i, column=6, value=lanc.get("valor", 0.0)).number_format = FMT_BRL
        ws.cell(row=i, column=7, value=lanc.get("classificacao", ""))
        ws.cell(row=i, column=8, value=lanc.get("categoria", ""))
        ws.cell(row=i, column=9, value=lanc.get("confianca", 0.0)).number_format = '0.00'
        ws.cell(row=i, column=10, value=origem)
        ws.cell(row=i, column=11, value=lanc.get("parcela_atual"))
        ws.cell(row=i, column=12, value=lanc.get("total_parcelas"))

        # Cor de fundo por tipo
        fill = None
        if tipo == "pagamento":
            fill = FILL_VERDE
        elif tipo == "encargo":
            fill = FILL_VERMELHO
        elif tipo == "parcelada":
            fill = FILL_AMARELO

        if fill:
            for col in range(1, 13):
                ws.cell(row=i, column=col).fill = fill

        # Coluna PF/PJ amarela para pendentes (sobrepoe o fill do tipo)
        if is_pendente:
            ws.cell(row=i, column=7).fill = FILL_AMARELO
            ws.cell(row=i, column=8).fill = FILL_AMARELO

        # Linhas alternadas cinza (se nao tem fill de tipo)
        if not fill and i % 2 == 0:
            for col in range(1, 13):
                ws.cell(row=i, column=col).fill = FILL_CINZA

    _aplicar_estilos(ws, num_colunas=12, num_linhas=len(lancamentos) + 1)
    ws.sheet_properties.tabColor = CORES["header"]


def _aba_parcelamentos(
    wb: Workbook, lancamentos: list[dict], mes_referencia: str
) -> None:
    """Aba 'Parcelamentos' — so parceladas com parcelas restantes e ultima fatura."""
    ws = wb.create_sheet(title="Parcelamentos")

    headers = [
        "Data Compra", "Descrição", "Cidade", "Valor Parcela (R$)",
        "Parcela Atual", "Total Parcelas", "Parcelas Restantes",
        "Última Fatura Prevista", "PF/PJ", "Categoria",
    ]
    _escrever_header(ws, headers)

    for i, lanc in enumerate(lancamentos, start=2):
        parcelas_rest = lanc.get("parcelas_restantes") or 0
        ultima = _calcular_ultima_fatura(mes_referencia, parcelas_rest)

        ws.cell(row=i, column=1, value=lanc.get("data", ""))
        ws.cell(row=i, column=2, value=lanc.get("descricao", ""))
        ws.cell(row=i, column=3, value=lanc.get("cidade", ""))
        ws.cell(row=i, column=4, value=lanc.get("valor", 0.0)).number_format = FMT_BRL
        ws.cell(row=i, column=5, value=lanc.get("parcela_atual"))
        ws.cell(row=i, column=6, value=lanc.get("total_parcelas"))
        ws.cell(row=i, column=7, value=parcelas_rest)
        ws.cell(row=i, column=8, value=ultima)
        ws.cell(row=i, column=9, value=lanc.get("classificacao", ""))
        ws.cell(row=i, column=10, value=lanc.get("categoria", ""))

        # Linhas alternadas
        if i % 2 == 0:
            for col in range(1, 11):
                ws.cell(row=i, column=col).fill = FILL_CINZA

    _aplicar_estilos(ws, num_colunas=10, num_linhas=len(lancamentos) + 1)
    ws.sheet_properties.tabColor = CORES["amarelo"]


def _aba_projecao(
    wb: Workbook, lancamentos: list[dict], mes_referencia: str
) -> None:
    """Aba 'Projecao Faturas' — parcelas futuras agrupadas por mes.
    Linhas: estabelecimentos. Colunas: meses futuros.
    """
    ws = wb.create_sheet(title="Projecao Faturas")

    if not lancamentos or not mes_referencia:
        ws.cell(row=1, column=1, value="Sem parcelamentos para projetar.")
        return

    # Agrupar por descricao
    grupos = defaultdict(list)
    for lanc in lancamentos:
        desc = lanc.get("descricao", "Sem descricao")
        grupos[desc].append(lanc)

    # Calcular meses futuros necessarios
    max_restantes = 0
    for lanc in lancamentos:
        rest = lanc.get("parcelas_restantes") or 0
        if rest > max_restantes:
            max_restantes = rest

    if max_restantes == 0:
        ws.cell(row=1, column=1, value="Todas as parcelas encerram nesta fatura.")
        return

    # Gerar lista de meses futuros
    meses_futuros = []
    ano, mes = int(mes_referencia[:4]), int(mes_referencia[5:7])
    for _ in range(max_restantes):
        mes += 1
        if mes > 12:
            mes = 1
            ano += 1
        meses_futuros.append(f"{ano:04d}-{mes:02d}")

    # Header
    headers = ["Descrição", "Valor Parcela (R$)"] + meses_futuros + ["Total Futuro (R$)"]
    _escrever_header(ws, headers)

    row = 2
    total_geral_por_mes = [0.0] * len(meses_futuros)
    total_geral_futuro = 0.0

    for desc in sorted(grupos.keys()):
        lancs = grupos[desc]
        # Somar valores de lancamentos com mesma descricao
        valor_parcela = sum(l.get("valor", 0.0) for l in lancs)
        max_rest = max((l.get("parcelas_restantes") or 0) for l in lancs)

        ws.cell(row=row, column=1, value=desc)
        ws.cell(row=row, column=2, value=valor_parcela).number_format = FMT_BRL

        total_futuro_linha = 0.0
        for j in range(len(meses_futuros)):
            if j < max_rest:
                ws.cell(row=row, column=3 + j, value=valor_parcela).number_format = FMT_BRL
                total_geral_por_mes[j] += valor_parcela
                total_futuro_linha += valor_parcela
            else:
                ws.cell(row=row, column=3 + j, value="")

        col_total = 3 + len(meses_futuros)
        ws.cell(row=row, column=col_total, value=total_futuro_linha).number_format = FMT_BRL
        total_geral_futuro += total_futuro_linha

        if row % 2 == 0:
            for col in range(1, col_total + 1):
                ws.cell(row=row, column=col).fill = FILL_CINZA

        row += 1

    # Linha de totais
    ws.cell(row=row, column=1, value="TOTAL").font = FONT_BOLD
    ws.cell(row=row, column=2, value="")
    for j, total_mes in enumerate(total_geral_por_mes):
        cell = ws.cell(row=row, column=3 + j, value=total_mes)
        cell.number_format = FMT_BRL
        cell.font = FONT_BOLD
    col_total = 3 + len(meses_futuros)
    cell = ws.cell(row=row, column=col_total, value=total_geral_futuro)
    cell.number_format = FMT_BRL
    cell.font = FONT_BOLD

    for col in range(1, col_total + 1):
        ws.cell(row=row, column=col).fill = FILL_PREVIA

    _aplicar_estilos(ws, num_colunas=col_total, num_linhas=row)
    ws.sheet_properties.tabColor = CORES["previa"]


def _aba_resumo(wb: Workbook, resumo: dict, stats: dict, banco: str) -> None:
    """Aba 'Resumo Fatura' — visao executiva."""
    ws = wb.create_sheet(title="Resumo Fatura")

    dados = [
        ("Banco", banco),
        ("Mês Referência", resumo.get("mes_referencia", "")),
        ("Vencimento", resumo.get("vencimento", "")),
        ("", ""),
        ("Saldo Fatura Anterior", resumo.get("saldo_anterior", 0.0)),
        ("Pagamentos/Créditos", resumo.get("pagamentos", 0.0)),
        ("Tarifas/Encargos/Multas", resumo.get("encargos", 0.0)),
        ("Total da Fatura", resumo.get("total_fatura", 0.0)),
        ("Saldo Parcelado Futuro", resumo.get("saldo_parcelado_futuro", 0.0)),
        ("", ""),
        ("Classificação", ""),
        ("Total Lançamentos", stats.get("total", 0)),
        ("Classificados", stats.get("classificados", 0)),
        ("Pendentes", stats.get("pendentes", 0)),
        ("Via Histórico", stats.get("por_origem", {}).get("historico", 0)),
        ("Via Recorrência", stats.get("por_origem", {}).get("recorrencia", 0)),
    ]

    headers = ["Campo", "Valor"]
    _escrever_header(ws, headers)

    for i, (campo, valor) in enumerate(dados, start=2):
        ws.cell(row=i, column=1, value=campo).font = FONT_BOLD if campo else FONT_NORMAL
        cell_val = ws.cell(row=i, column=2, value=valor)
        if isinstance(valor, float):
            cell_val.number_format = FMT_BRL

        # Destaque para total
        if campo == "Total da Fatura":
            ws.cell(row=i, column=1).fill = FILL_PREVIA
            ws.cell(row=i, column=2).fill = FILL_PREVIA

        # Destaque pendentes
        if campo == "Pendentes" and valor > 0:
            ws.cell(row=i, column=1).fill = FILL_AMARELO
            ws.cell(row=i, column=2).fill = FILL_AMARELO

    _aplicar_estilos(ws, num_colunas=2, num_linhas=len(dados) + 1)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.sheet_properties.tabColor = CORES["header"]


def _aba_pf_pj(
    wb: Workbook,
    lancamentos: list[dict],
    apenas_se_completo: bool = True,
) -> bool:
    """Gera aba 'PF vs PJ' apenas se nao houver pendentes (ou se forcado).
    Retorna True se gerou, False se pulou.
    """
    pendentes = [l for l in lancamentos if l.get("origem") == "pendente"]
    if apenas_se_completo and pendentes:
        return False

    ws = wb.create_sheet(title="PF vs PJ")

    # Agrupar por classificacao e categoria (so compras, exclui pagamento/encargo)
    totais = {"PF": defaultdict(float), "PJ": defaultdict(float)}
    for lanc in lancamentos:
        classif = lanc.get("classificacao", "")
        if classif not in ("PF", "PJ"):
            continue
        if lanc.get("tipo") not in TIPOS_PF_PJ:
            continue
        categoria = lanc.get("categoria", "Outros")
        totais[classif][categoria] += lanc.get("valor", 0.0)

    headers = ["Classificação", "Categoria", "Total (R$)", "% do Total"]
    _escrever_header(ws, headers)

    total_geral = sum(
        sum(cats.values()) for cats in totais.values()
    )

    row = 2
    for classif in ("PJ", "PF"):
        cats = totais[classif]
        if not cats:
            continue

        subtotal = sum(cats.values())
        fill_sub = FILL_SUBTOTAL_PJ if classif == "PJ" else FILL_SUBTOTAL_PF
        font_sub = FONT_SUBTOTAL_PJ if classif == "PJ" else FONT_SUBTOTAL_PF

        for categoria in sorted(cats.keys()):
            valor = cats[categoria]
            pct = (valor / total_geral * 100) if total_geral else 0.0

            ws.cell(row=row, column=1, value=classif)
            ws.cell(row=row, column=2, value=categoria)
            ws.cell(row=row, column=3, value=valor).number_format = FMT_BRL
            ws.cell(row=row, column=4, value=f"{pct:.1f}%")

            fill = FILL_AZUL if classif == "PF" else FILL_VERDE
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill

            row += 1

        # Subtotal
        pct_sub = (subtotal / total_geral * 100) if total_geral else 0.0
        ws.cell(row=row, column=1, value="").fill = fill_sub
        ws.cell(row=row, column=2, value=f"Subtotal {classif}").font = font_sub
        ws.cell(row=row, column=2).fill = fill_sub
        ws.cell(row=row, column=3, value=subtotal).number_format = FMT_BRL
        ws.cell(row=row, column=3).font = font_sub
        ws.cell(row=row, column=3).fill = fill_sub
        ws.cell(row=row, column=4, value=f"{pct_sub:.1f}%").font = font_sub
        ws.cell(row=row, column=4).fill = fill_sub
        row += 1

    # Total geral
    ws.cell(row=row, column=2, value="TOTAL GERAL").font = FONT_BOLD
    ws.cell(row=row, column=3, value=total_geral).number_format = FMT_BRL
    ws.cell(row=row, column=3).font = FONT_BOLD
    ws.cell(row=row, column=4, value="100.0%").font = FONT_BOLD
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = FILL_PREVIA

    _aplicar_estilos(ws, num_colunas=4, num_linhas=row)
    ws.sheet_properties.tabColor = CORES["azul_pj"]
    return True


def _aba_relatorio_conselho(
    wb: Workbook,
    lancamentos: list[dict],
    resumo: dict,
    apenas_se_completo: bool = True,
) -> bool:
    """Gera aba 'Relatorio Conselho' apenas se nao houver pendentes (ou se forcado).
    Retorna True se gerou, False se pulou.
    """
    pendentes = [l for l in lancamentos if l.get("origem") == "pendente"]
    if apenas_se_completo and pendentes:
        return False

    ws = wb.create_sheet(title="Relatorio Conselho")

    # Cabecalho do relatorio
    ws.merge_cells("A1:D1")
    cell_titulo = ws.cell(row=1, column=1, value="Relatório de Prestação de Contas — Cartão Corporativo")
    cell_titulo.font = Font(bold=True, size=14, color=CORES["header"])
    cell_titulo.alignment = ALIGN_CENTER

    ws.merge_cells("A2:D2")
    ws.cell(
        row=2, column=1,
        value=f"Referência: {resumo.get('mes_referencia', '')} | Vencimento: {resumo.get('vencimento', '')}",
    ).alignment = ALIGN_CENTER

    # Resumo financeiro
    row = 4
    ws.cell(row=row, column=1, value="Resumo Financeiro").font = Font(bold=True, size=12)
    row += 1

    itens_resumo = [
        ("Total da Fatura", resumo.get("total_fatura", 0.0)),
        ("Saldo Anterior", resumo.get("saldo_anterior", 0.0)),
        ("Pagamentos/Créditos", resumo.get("pagamentos", 0.0)),
        ("Encargos", resumo.get("encargos", 0.0)),
        ("Saldo Parcelado Futuro", resumo.get("saldo_parcelado_futuro", 0.0)),
    ]
    for campo, valor in itens_resumo:
        ws.cell(row=row, column=1, value=campo).font = FONT_BOLD
        ws.cell(row=row, column=2, value=valor).number_format = FMT_BRL
        row += 1

    # Distribuicao PF vs PJ
    row += 1
    ws.cell(row=row, column=1, value="Distribuição PF vs PJ").font = Font(bold=True, size=12)
    row += 1

    total_pf = sum(l.get("valor", 0.0) for l in lancamentos if l.get("classificacao") == "PF" and l.get("tipo") in TIPOS_PF_PJ)
    total_pj = sum(l.get("valor", 0.0) for l in lancamentos if l.get("classificacao") == "PJ" and l.get("tipo") in TIPOS_PF_PJ)
    total_class = total_pf + total_pj

    pct_pf = (total_pf / total_class * 100) if total_class else 0.0
    pct_pj = (total_pj / total_class * 100) if total_class else 0.0

    ws.cell(row=row, column=1, value="Pessoa Física (PF)").font = FONT_BOLD
    ws.cell(row=row, column=2, value=total_pf).number_format = FMT_BRL
    ws.cell(row=row, column=3, value=f"{pct_pf:.1f}%")
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = FILL_AZUL
    row += 1

    ws.cell(row=row, column=1, value="Pessoa Jurídica (PJ)").font = FONT_BOLD
    ws.cell(row=row, column=2, value=total_pj).number_format = FMT_BRL
    ws.cell(row=row, column=3, value=f"{pct_pj:.1f}%")
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = FILL_VERDE
    row += 1

    # Detalhamento PJ por categoria
    row += 1
    ws.cell(row=row, column=1, value="Detalhamento PJ por Categoria").font = Font(bold=True, size=12)
    row += 1

    cats_pj = defaultdict(float)
    for lanc in lancamentos:
        if lanc.get("classificacao") == "PJ" and lanc.get("tipo") in TIPOS_PF_PJ:
            cats_pj[lanc.get("categoria", "Outros")] += lanc.get("valor", 0.0)

    _escrever_header_em_row(ws, row, ["Categoria", "Valor (R$)", "% PJ"])
    row += 1

    for cat in sorted(cats_pj.keys()):
        val = cats_pj[cat]
        pct = (val / total_pj * 100) if total_pj else 0.0
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=val).number_format = FMT_BRL
        ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        if row % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = FILL_CINZA
        row += 1

    _aplicar_estilos(ws, num_colunas=4, num_linhas=row - 1)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.sheet_properties.tabColor = CORES["verde_pj"]
    return True


# ===========================================================================
# EXTRATOS BANCARIOS
# ===========================================================================

FONT_RISCADO = Font(size=11, color="999999", strikethrough=True)
FILL_CINZA_ESTORNO = PatternFill("solid", fgColor="E0E0E0")
FILL_ALERTA = PatternFill("solid", fgColor="FCE4D6")


def gerar_excel_extrato(
    extrato: dict,
    classificacao: dict,
    config: dict,
    caminho_saida: str,
) -> str:
    """Entry point para Excel de extrato bancario. Retorna caminho final."""
    wb = Workbook()

    resumo = extrato.get("resumo", {})
    mes_referencia = extrato.get("mes_referencia", "")

    # Unificar lancamentos: classificados + pendentes + nao_efetivados (ordem original)
    todos = (
        classificacao.get("classificados", [])
        + classificacao.get("pendentes", [])
        + classificacao.get("nao_efetivados", [])
    )
    todos.sort(key=lambda l: l.get("data", ""))

    stats = classificacao.get("stats", {})

    # Aba 1: Lancamentos
    _aba_lancamentos_extrato(wb, todos, extrato.get("banco", "BB"))

    # Aba 2: Resumo
    _aba_resumo_extrato(wb, resumo, stats)

    # Aba 3: Por Categoria
    _aba_por_categoria(wb, todos)

    # Aba 4: Fluxo de Caixa
    _aba_fluxo_caixa(wb, todos, resumo.get("saldo_inicial", 0.0))

    # Remover sheet default vazia
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # Montar nome do arquivo
    meses_config = config.get("meses", {})
    titular = config.get("titular", "Hugo")
    banco = extrato.get("banco", "BB")

    if mes_referencia and len(mes_referencia) >= 7:
        ano = mes_referencia[:4]
        mes_num = int(mes_referencia[5:7])
        mes_nome = meses_config.get(mes_num, f"{mes_num:02d}")
        nome_arquivo = f"{titular} - Extrato - {banco} - {mes_nome}.{ano}.xlsx"
        subpasta = f"{ano}/{mes_nome}.{ano}"
    else:
        nome_arquivo = f"{titular} - Extrato - {banco}.xlsx"
        subpasta = ""

    caminho_final = Path(caminho_saida) / subpasta / nome_arquivo
    caminho_final.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(caminho_final))
    logger.info("Excel extrato salvo em %s", caminho_final)
    return str(caminho_final)


def _aba_lancamentos_extrato(
    wb: Workbook,
    lancamentos: list[dict],
    banco: str,
) -> None:
    """Aba 'Lancamentos' — todos os lancamentos do extrato.

    Nao efetivados em cinza riscado, entradas verde, saidas vermelho,
    pendentes com Categoria/Classificacao amarelo.
    """
    ws = wb.create_sheet(title=f"Lancamentos {banco}")

    headers = [
        "Data", "Descrição", "Detalhes", "Valor (R$)", "Tipo",
        "Categoria", "Classificação", "Efetivado",
    ]
    _escrever_header(ws, headers)

    for i, lanc in enumerate(lancamentos, start=2):
        efetivado = lanc.get("efetivado", True)
        origem = lanc.get("origem", "pendente")
        is_pendente = origem == "pendente"
        valor = lanc.get("valor", 0.0)

        ws.cell(row=i, column=1, value=lanc.get("data", ""))
        ws.cell(row=i, column=2, value=lanc.get("descricao", ""))
        ws.cell(row=i, column=3, value=lanc.get("detalhes", ""))
        ws.cell(row=i, column=4, value=valor).number_format = FMT_BRL
        ws.cell(row=i, column=5, value=lanc.get("tipo", ""))
        ws.cell(row=i, column=6, value=lanc.get("categoria", ""))
        ws.cell(row=i, column=7, value=lanc.get("classificacao", ""))
        ws.cell(row=i, column=8, value="Sim" if efetivado else "Não")

        if not efetivado:
            # Cinza + riscado
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = FILL_CINZA_ESTORNO
                ws.cell(row=i, column=col).font = FONT_RISCADO
        elif valor > 0:
            # Entrada: verde
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = FILL_VERDE
        elif valor < 0:
            # Saida: vermelho
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = FILL_VERMELHO

        # Pendentes: categoria/classificacao amarelo (sobrepoe)
        if is_pendente and efetivado:
            ws.cell(row=i, column=6).fill = FILL_AMARELO
            ws.cell(row=i, column=7).fill = FILL_AMARELO

    _aplicar_estilos(ws, num_colunas=8, num_linhas=len(lancamentos) + 1)
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.sheet_properties.tabColor = CORES["header"]


def _aba_resumo_extrato(
    wb: Workbook,
    resumo: dict,
    stats: dict,
) -> None:
    """Aba 'Resumo' — visao executiva com saldos, totais e alertas."""
    ws = wb.create_sheet(title="Resumo")

    dados = [
        ("Banco", resumo.get("banco", "BB")),
        ("Conta", resumo.get("conta", "corrente")),
        ("Mês Referência", resumo.get("mes_referencia", "")),
        ("", ""),
        ("Saldo Inicial", resumo.get("saldo_inicial", 0.0)),
        ("Saldo Final", resumo.get("saldo_final", 0.0)),
        ("", ""),
        ("Total Entradas Efetivadas", resumo.get("total_entradas", 0.0)),
        ("Total Saídas Efetivadas", resumo.get("total_saidas", 0.0)),
        ("Total Não Efetivado", resumo.get("total_estornados", 0.0)),
        ("", ""),
        ("Classificação", ""),
        ("Total Lançamentos", stats.get("total", 0)),
        ("Classificados", stats.get("classificados", 0)),
        ("Pendentes", stats.get("pendentes", 0)),
        ("Não Efetivados", stats.get("nao_efetivados", 0)),
        ("Via Histórico", stats.get("por_origem", {}).get("historico", 0)),
        ("Via Recorrência", stats.get("por_origem", {}).get("recorrencia", 0)),
    ]

    headers = ["Campo", "Valor"]
    _escrever_header(ws, headers)

    for i, (campo, valor) in enumerate(dados, start=2):
        ws.cell(row=i, column=1, value=campo).font = FONT_BOLD if campo else FONT_NORMAL
        cell_val = ws.cell(row=i, column=2, value=valor)
        if isinstance(valor, float):
            cell_val.number_format = FMT_BRL

        # Destaque saldo final
        if campo == "Saldo Final":
            fill = FILL_VERDE if valor >= 0 else FILL_VERMELHO
            ws.cell(row=i, column=1).fill = fill
            ws.cell(row=i, column=2).fill = fill

        # Destaque pendentes
        if campo == "Pendentes" and valor > 0:
            ws.cell(row=i, column=1).fill = FILL_AMARELO
            ws.cell(row=i, column=2).fill = FILL_AMARELO

        # Destaque nao efetivados
        if campo == "Total Não Efetivado" and valor > 0:
            ws.cell(row=i, column=1).fill = FILL_CINZA_ESTORNO
            ws.cell(row=i, column=2).fill = FILL_CINZA_ESTORNO

    # Alertas
    alertas = resumo.get("alertas", [])
    if alertas:
        row = len(dados) + 3
        ws.cell(row=row, column=1, value="ALERTAS").font = Font(bold=True, size=12, color="CC0000")
        row += 1
        for alerta in alertas:
            ws.cell(row=row, column=1, value=alerta)
            ws.cell(row=row, column=1).fill = FILL_ALERTA
            ws.cell(row=row, column=2).fill = FILL_ALERTA
            ws.cell(row=row, column=1).font = Font(size=11, color="CC0000")
            row += 1

        _aplicar_estilos(ws, num_colunas=2, num_linhas=row - 1)
    else:
        _aplicar_estilos(ws, num_colunas=2, num_linhas=len(dados) + 1)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.sheet_properties.tabColor = CORES["header"]


def _aba_por_categoria(
    wb: Workbook,
    lancamentos: list[dict],
) -> None:
    """Aba 'Por Categoria' — receitas e despesas agrupadas por categoria.

    Secoes: Receitas, Despesas PF, Despesas PJ com total/% /num lancamentos.
    """
    ws = wb.create_sheet(title="Por Categoria")

    # Agrupar lancamentos efetivados por secao
    receitas: dict[str, list[float]] = defaultdict(list)
    despesas_pf: dict[str, list[float]] = defaultdict(list)
    despesas_pj: dict[str, list[float]] = defaultdict(list)

    for lanc in lancamentos:
        if not lanc.get("efetivado", True):
            continue
        if lanc.get("origem") == "pendente":
            continue

        classif = lanc.get("classificacao", "")
        categoria = lanc.get("categoria", "") or "Sem Categoria"
        valor = lanc.get("valor", 0.0)

        if classif == "PF":
            despesas_pf[categoria].append(valor)
        elif classif == "PJ":
            despesas_pj[categoria].append(valor)
        elif valor > 0:
            receitas[categoria].append(valor)

    row = 1

    # --- Secao Receitas ---
    row = _escrever_secao_categoria(
        ws, row, "RECEITAS", receitas, FILL_VERDE, FONT_BOLD,
    )

    row += 1

    # --- Secao Despesas PF ---
    row = _escrever_secao_categoria(
        ws, row, "DESPESAS PF", despesas_pf, FILL_AZUL, FONT_BOLD,
    )

    row += 1

    # --- Secao Despesas PJ ---
    row = _escrever_secao_categoria(
        ws, row, "DESPESAS PJ", despesas_pj, FILL_VERDE, FONT_BOLD,
    )

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.sheet_properties.tabColor = CORES["azul_pj"]


def _escrever_secao_categoria(
    ws,
    row_inicio: int,
    titulo: str,
    dados: dict[str, list[float]],
    fill_secao: PatternFill,
    font_titulo: Font,
) -> int:
    """Escreve uma secao de categorias na aba Por Categoria.

    Retorna a proxima row disponivel.
    """
    row = row_inicio

    # Titulo da secao
    ws.cell(row=row, column=1, value=titulo).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = fill_secao
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = fill_secao
    row += 1

    if not dados:
        ws.cell(row=row, column=1, value="Nenhum lançamento classificado")
        ws.cell(row=row, column=1).font = Font(size=11, color="999999", italic=True)
        return row + 1

    # Header
    _escrever_header_em_row(ws, row, ["Categoria", "Total (R$)", "% do Total", "Nº Lançamentos"])
    row += 1

    total_secao = sum(sum(vals) for vals in dados.values())

    for categoria in sorted(dados.keys()):
        valores = dados[categoria]
        total_cat = sum(valores)
        pct = (total_cat / total_secao * 100) if total_secao else 0.0

        ws.cell(row=row, column=1, value=categoria)
        ws.cell(row=row, column=2, value=total_cat).number_format = FMT_BRL
        ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        ws.cell(row=row, column=4, value=len(valores))

        if row % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = FILL_CINZA

        row += 1

    # Subtotal
    ws.cell(row=row, column=1, value=f"Total {titulo}").font = FONT_BOLD
    ws.cell(row=row, column=2, value=total_secao).number_format = FMT_BRL
    ws.cell(row=row, column=2).font = FONT_BOLD
    ws.cell(row=row, column=3, value="100.0%").font = FONT_BOLD
    total_lancs = sum(len(vals) for vals in dados.values())
    ws.cell(row=row, column=4, value=total_lancs).font = FONT_BOLD
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = FILL_PREVIA

    return row + 1


def _aba_fluxo_caixa(
    wb: Workbook,
    lancamentos: list[dict],
    saldo_inicial: float,
) -> None:
    """Aba 'Fluxo de Caixa' — por dia, so efetivados.

    Data | Entradas do dia | Saidas do dia | Saldo Acumulado.
    """
    ws = wb.create_sheet(title="Fluxo de Caixa")

    headers = ["Data", "Entradas (R$)", "Saídas (R$)", "Saldo Acumulado (R$)"]
    _escrever_header(ws, headers)

    # Agrupar efetivados por dia
    por_dia: dict[str, dict] = defaultdict(lambda: {"entradas": 0.0, "saidas": 0.0})
    for lanc in lancamentos:
        if not lanc.get("efetivado", True):
            continue
        data = lanc.get("data", "")
        if not data:
            continue
        valor = lanc.get("valor", 0.0)
        if valor > 0:
            por_dia[data]["entradas"] += valor
        else:
            por_dia[data]["saidas"] += valor

    saldo = saldo_inicial
    row = 2

    # Linha de saldo inicial
    ws.cell(row=row, column=1, value="Saldo Inicial")
    ws.cell(row=row, column=2, value="")
    ws.cell(row=row, column=3, value="")
    ws.cell(row=row, column=4, value=saldo).number_format = FMT_BRL
    ws.cell(row=row, column=1).font = FONT_BOLD
    ws.cell(row=row, column=4).font = FONT_BOLD
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = FILL_PREVIA
    row += 1

    for data in sorted(por_dia.keys()):
        dia = por_dia[data]
        entradas = dia["entradas"]
        saidas = dia["saidas"]
        saldo += entradas + saidas

        ws.cell(row=row, column=1, value=data)
        ws.cell(row=row, column=2, value=entradas).number_format = FMT_BRL
        ws.cell(row=row, column=3, value=saidas).number_format = FMT_BRL
        cell_saldo = ws.cell(row=row, column=4, value=round(saldo, 2))
        cell_saldo.number_format = FMT_BRL

        # Entrada verde, saida vermelha
        if entradas > 0:
            ws.cell(row=row, column=2).fill = FILL_VERDE
        if saidas < 0:
            ws.cell(row=row, column=3).fill = FILL_VERMELHO

        # Saldo negativo em vermelho
        if saldo < 0:
            cell_saldo.fill = FILL_VERMELHO
            cell_saldo.font = Font(bold=True, size=11, color="CC0000")

        row += 1

    _aplicar_estilos(ws, num_colunas=4, num_linhas=row - 1)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.sheet_properties.tabColor = CORES["verde_l"]


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _escrever_header(ws, headers: list[str], row: int = 1) -> None:
    """Escreve linha de header com estilo padrao."""
    for col, texto in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=texto)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _escrever_header_em_row(ws, row: int, headers: list[str]) -> None:
    """Escreve header em uma row especifica (para abas com conteudo misto)."""
    _escrever_header(ws, headers, row=row)


def _aplicar_estilos(ws, num_colunas: int, num_linhas: int, header_row: int = 1) -> None:
    """Aplica bordas, alinhamento e largura de colunas."""
    # Bordas em todas as celulas com dados
    for row in range(1, num_linhas + 1):
        for col in range(1, num_colunas + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if cell.alignment == Alignment():
                cell.alignment = ALIGN_LEFT

    # Auto-width baseado no header
    for col in range(1, num_colunas + 1):
        header_val = ws.cell(row=header_row, column=col).value
        width = max(len(str(header_val or "")) + 4, 12)
        ws.column_dimensions[get_column_letter(col)].width = width

    # Congelar painel abaixo do header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _calcular_ultima_fatura(mes_referencia: str, parcelas_restantes: int) -> str:
    """Retorna 'YYYY-MM' da ultima fatura prevista para um parcelamento."""
    if not mes_referencia or parcelas_restantes <= 0:
        return mes_referencia or ""

    ano = int(mes_referencia[:4])
    mes = int(mes_referencia[5:7])

    mes += parcelas_restantes
    while mes > 12:
        mes -= 12
        ano += 1

    return f"{ano:04d}-{mes:02d}"


# ===========================================================================
# CONTAS A PAGAR — Projeção 12 meses
# ===========================================================================

FILL_ATRASADO = PatternFill("solid", fgColor="F4CCCC")
FILL_TIPO_HEADER = PatternFill("solid", fgColor="D9E2F3")

TIPO_LABELS = {
    "fatura_cartao": "Faturas Cartão",
    "parcela": "Parcelas",
    "recorrencia": "Recorrências",
    "fixo_spm": "Fixo SPM",
    "debito_nao_efetivado": "Débitos Não Efetivados",
    "manual": "Manual",
}


def gerar_excel_contas_a_pagar(
    projecao: dict,
    config: dict,
    caminho_saida: str,
) -> str:
    """Entry point para Excel de contas a pagar. Retorna caminho final."""
    wb = Workbook()

    compromissos = projecao.get("compromissos", [])
    resumo_por_mes = projecao.get("resumo_por_mes", {})
    mes_referencia = projecao.get("mes_referencia", "")

    _aba_visao_por_data(wb, compromissos)
    _aba_visao_por_valor(wb, compromissos)
    _aba_visao_por_tipo(wb, compromissos)
    _aba_projecao_12_meses(wb, projecao, config)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # Montar caminho
    titular = config.get("titular", "Hugo")
    nome_arquivo = f"{titular} - Contas a Pagar - Projecao 12 Meses.xlsx"

    if mes_referencia and len(mes_referencia) >= 7:
        ano = mes_referencia[:4]
        caminho_final = Path(caminho_saida) / ano / nome_arquivo
    else:
        caminho_final = Path(caminho_saida) / nome_arquivo

    caminho_final.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(caminho_final))
    logger.info("Excel contas a pagar salvo em %s", caminho_final)
    return str(caminho_final)


def _aba_visao_por_data(wb: Workbook, compromissos: list[dict]) -> None:
    """Aba 'Visão por Data' — compromissos ordenados por data de vencimento."""
    ws = wb.create_sheet(title="Visão por Data")

    headers = [
        "Mês", "Data Vencimento", "Descrição", "Valor (R$)",
        "Tipo", "Categoria", "Pago por", "Status",
    ]
    _escrever_header(ws, headers)

    ordenados = sorted(compromissos, key=lambda c: c.get("data_vencimento", ""))
    _preencher_linhas_compromissos(ws, ordenados, start_row=2)

    _aplicar_estilos(ws, num_colunas=8, num_linhas=len(ordenados) + 1)
    ws.column_dimensions["C"].width = 40
    ws.sheet_properties.tabColor = CORES["header"]


def _aba_visao_por_valor(wb: Workbook, compromissos: list[dict]) -> None:
    """Aba 'Visão por Valor' — compromissos ordenados por valor decrescente."""
    ws = wb.create_sheet(title="Visão por Valor")

    headers = [
        "Mês", "Data Vencimento", "Descrição", "Valor (R$)",
        "Tipo", "Categoria", "Pago por", "Status",
    ]
    _escrever_header(ws, headers)

    ordenados = sorted(compromissos, key=lambda c: c.get("valor", 0.0), reverse=True)
    _preencher_linhas_compromissos(ws, ordenados, start_row=2)

    _aplicar_estilos(ws, num_colunas=8, num_linhas=len(ordenados) + 1)
    ws.column_dimensions["C"].width = 40
    ws.sheet_properties.tabColor = CORES["vermelho_l"]


def _aba_visao_por_tipo(wb: Workbook, compromissos: list[dict]) -> None:
    """Aba 'Visão por Tipo' — agrupado por tipo com subtotais."""
    ws = wb.create_sheet(title="Visão por Tipo")

    row = 1
    total_geral = 0.0

    tipos_ordem = ["fatura_cartao", "parcela", "recorrencia", "fixo_spm", "debito_nao_efetivado", "manual"]

    for tipo in tipos_ordem:
        grupo = [c for c in compromissos if c.get("tipo") == tipo]
        if not grupo:
            continue

        # Header da secao
        label = TIPO_LABELS.get(tipo, tipo)
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=12)
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = FILL_TIPO_HEADER
        row += 1

        _escrever_header(ws, [
            "Mês", "Data Vencimento", "Descrição", "Valor (R$)",
            "Tipo", "Categoria", "Pago por", "Status",
        ], row=row)
        row += 1

        grupo_sorted = sorted(grupo, key=lambda c: c.get("data_vencimento", ""))
        for c in grupo_sorted:
            _escrever_linha_compromisso(ws, row, c)
            row += 1

        # Subtotal
        subtotal = sum(c.get("valor", 0.0) for c in grupo)
        total_geral += subtotal
        ws.cell(row=row, column=3, value=f"Subtotal {label}").font = FONT_BOLD
        ws.cell(row=row, column=4, value=subtotal).number_format = FMT_BRL
        ws.cell(row=row, column=4).font = FONT_BOLD
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = FILL_PREVIA
        row += 2

    # Total geral
    ws.cell(row=row, column=3, value="TOTAL GERAL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=4, value=total_geral).number_format = FMT_BRL
    ws.cell(row=row, column=4).font = Font(bold=True, size=12)
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = FILL_PREVIA

    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 16
    ws.sheet_properties.tabColor = CORES["azul_pj"]


def _aba_projecao_12_meses(wb: Workbook, projecao: dict, config: dict) -> None:
    """Aba 'Projeção 12 Meses' — uma coluna por mês, uma linha por categoria."""
    ws = wb.create_sheet(title="Projeção 12 Meses")

    resumo_por_mes = projecao.get("resumo_por_mes", {})
    meses = sorted(resumo_por_mes.keys())
    meses_config = config.get("meses", {})

    if not meses:
        ws.cell(row=1, column=1, value="Sem dados para projetar.")
        return

    # Coletar todas as categorias
    todas_categorias = set()
    for mes in meses:
        todas_categorias.update(resumo_por_mes[mes].get("por_categoria", {}).keys())
    categorias = sorted(todas_categorias)

    # Header: Categoria | Abr/26 | Mai/26 | ... | Total
    headers = ["Categoria"]
    for mes in meses:
        ano = mes[:4]
        mes_num = int(mes[5:7])
        nome_mes = meses_config.get(mes_num, f"{mes_num:02d}")
        headers.append(f"{nome_mes[:3]}/{ano[2:]}")
    headers.append("Total (R$)")
    _escrever_header(ws, headers)

    row = 2

    # Uma linha por categoria
    for cat in categorias:
        ws.cell(row=row, column=1, value=cat)
        total_cat = 0.0

        for j, mes in enumerate(meses):
            val = resumo_por_mes[mes].get("por_categoria", {}).get(cat, 0.0)
            cell = ws.cell(row=row, column=2 + j, value=val if val else "")
            if val:
                cell.number_format = FMT_BRL
                total_cat += val

        ws.cell(row=row, column=2 + len(meses), value=total_cat).number_format = FMT_BRL

        if row % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = FILL_CINZA

        row += 1

    # Linha de total por mes
    ws.cell(row=row, column=1, value="TOTAL MÊS").font = FONT_BOLD
    total_geral = 0.0
    for j, mes in enumerate(meses):
        val = resumo_por_mes[mes].get("total", 0.0)
        cell = ws.cell(row=row, column=2 + j, value=val)
        cell.number_format = FMT_BRL
        cell.font = FONT_BOLD
        total_geral += val
    ws.cell(row=row, column=2 + len(meses), value=total_geral).number_format = FMT_BRL
    ws.cell(row=row, column=2 + len(meses)).font = FONT_BOLD
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).fill = FILL_PREVIA
    row += 1

    # Linha Hugo vs SPM
    for pagador, fill in [("Hugo", FILL_AZUL), ("SPM", FILL_VERDE)]:
        ws.cell(row=row, column=1, value=pagador).font = FONT_BOLD
        total_pag = 0.0
        for j, mes in enumerate(meses):
            val = resumo_por_mes[mes].get("por_pagador", {}).get(pagador, 0.0)
            cell = ws.cell(row=row, column=2 + j, value=val if val else "")
            if val:
                cell.number_format = FMT_BRL
                total_pag += val
        ws.cell(row=row, column=2 + len(meses), value=total_pag).number_format = FMT_BRL
        ws.cell(row=row, column=2 + len(meses)).font = FONT_BOLD
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = fill
        row += 1

    _aplicar_estilos(ws, num_colunas=len(headers), num_linhas=row - 1)
    ws.column_dimensions["A"].width = 28
    ws.sheet_properties.tabColor = CORES["previa"]


# ---------------------------------------------------------------------------
# Helpers contas a pagar
# ---------------------------------------------------------------------------

def _preencher_linhas_compromissos(ws, compromissos: list[dict], start_row: int) -> None:
    """Preenche linhas de compromissos com cores por status."""
    for i, c in enumerate(compromissos, start=start_row):
        _escrever_linha_compromisso(ws, i, c)


def _escrever_linha_compromisso(ws, row: int, c: dict) -> None:
    """Escreve uma linha de compromisso com cor por status/pagador."""
    ws.cell(row=row, column=1, value=c.get("mes", ""))
    ws.cell(row=row, column=2, value=c.get("data_vencimento", ""))
    ws.cell(row=row, column=3, value=c.get("descricao", ""))
    ws.cell(row=row, column=4, value=c.get("valor", 0.0)).number_format = FMT_BRL
    ws.cell(row=row, column=5, value=TIPO_LABELS.get(c.get("tipo", ""), c.get("tipo", "")))
    ws.cell(row=row, column=6, value=c.get("categoria", ""))
    ws.cell(row=row, column=7, value=c.get("pago_por", ""))
    ws.cell(row=row, column=8, value=c.get("status", ""))

    status = c.get("status", "")
    pago_por = c.get("pago_por", "")

    if status == "atrasado":
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = FILL_ATRASADO
    elif pago_por == "SPM":
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = FILL_VERDE
    elif row % 2 == 0:
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = FILL_CINZA
