"""
Parser do XLSX de orçamento mensal da SPM.

Formato (derivado do arquivo real SPM-Orcamento-04.2026.xlsx):
  - Sheet principal: "Contas a Pagar"
  - Linha 1: cabeçalho global do mês
  - 6 seções com subcabeçalho em coluna C:
      "DESPESAS FIXAS"
      "TRIBUTOS"
      "DESPESAS VARIÁVEIS"
      "COMISSÕES"
      "Pagamentos de Valores Variáveis"
      "DESPESAS DE PROFISSIONAIS"
  - Após cada subcabeçalho vêm 1-2 linhas de formato + cabeçalho de colunas,
    depois as linhas de dados até o próximo subcabeçalho ou EOF.
"""
from __future__ import annotations

import re
import logging
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import List, Optional, Tuple, Union

import openpyxl

logger = logging.getLogger(__name__)


class NaturezaOrcamento(str, Enum):
    DESPESA_FIXA = "DESPESA_FIXA"
    TRIBUTO = "TRIBUTO"
    SALARIO_VARIAVEL = "SALARIO_VARIAVEL"
    COMISSAO = "COMISSAO"
    VALOR_VARIAVEL = "VALOR_VARIAVEL"
    DESPESA_PROFISSIONAIS = "DESPESA_PROFISSIONAIS"


# Padrões regex para matching dos subcabeçalhos (case-insensitive, tolerante a espaços)
_SUBHEADERS: List[Tuple[re.Pattern, NaturezaOrcamento]] = [
    (re.compile(r"^\s*DESPESAS\s+FIXAS\s*$", re.I), NaturezaOrcamento.DESPESA_FIXA),
    (re.compile(r"^\s*TRIBUTOS\s*$", re.I), NaturezaOrcamento.TRIBUTO),
    (re.compile(r"^\s*DESPESAS\s+VARI[AÁ]VEIS\s*$", re.I), NaturezaOrcamento.SALARIO_VARIAVEL),
    (re.compile(r"^\s*COMISS[OÕ]ES\s*$", re.I), NaturezaOrcamento.COMISSAO),
    (re.compile(r"^\s*Pagamentos\s+de\s+Valores\s+Vari[aá]veis\s*$", re.I), NaturezaOrcamento.VALOR_VARIAVEL),
    (re.compile(r"^\s*DESPESAS\s+DE\s+PROFISSIONAIS\s*$", re.I), NaturezaOrcamento.DESPESA_PROFISSIONAIS),
]


# Subcabeçalhos que, no XLSX real do Hugo, aparecem APÓS a seção DESPESA_PROFISSIONAIS
# e marcam blocos que NÃO pertencem a profissionais (investimentos pessoais, cartões,
# empréstimos Bradesco, despesas pessoais, linhas-resumo). Usados como sentinelas de fim.
_SUBHEADERS_FIM_PROFISSIONAIS: List[re.Pattern] = [
    re.compile(r"^\s*INVESTIMENTOS\s*$", re.I),
    re.compile(r"^\s*DESPESAS\s+PESSOAIS(?:\s+DR\s+HUGO)?\s*$", re.I),
    re.compile(r"^\s*CART[OÕ]ES\s+DE\s+CR[ÉE]DITO\s*$", re.I),
    re.compile(r"^\s*EMPR[ÉE]STIMOS\s*$", re.I),
    re.compile(r"^\s*FINANCEIRO\s+PESSOAL\s*$", re.I),
    re.compile(r"^\s*N[ÃA]O\s+IDENTIFICADO\s*$", re.I),
    re.compile(r"^\s*RESULTADO\s+BRUTO\s*$", re.I),
    re.compile(r"^\s*MARGEM\s+BRUTA\s*$", re.I),
    re.compile(r"^\s*VALOR\s+TOTAL\s+GERAL\s*$", re.I),
    re.compile(r"^\s*Faturamento\s*$", re.I),
    re.compile(r"^\s*Despesas\s*$", re.I),
    re.compile(r"^\s*Saldo\s*$", re.I),
    re.compile(r"^\s*TOTAL\s*$", re.I),
]


def _eh_fim_secao_profissionais(ws, row_idx: int) -> bool:
    """Retorna True se a linha marca o fim da seção DESPESA_PROFISSIONAIS.

    Verifica colunas B-E (2..5) buscando qualquer sentinela textual conhecida.
    """
    for col_idx in range(2, 6):
        cell = ws.cell(row=row_idx, column=col_idx).value
        if not isinstance(cell, str):
            continue
        texto = cell.strip()
        if not texto:
            continue
        for pattern in _SUBHEADERS_FIM_PROFISSIONAIS:
            if pattern.match(texto):
                return True
    return False


@dataclass
class OrcamentoLinhaParsed:
    natureza: NaturezaOrcamento
    titular_razao_social: str
    empresa_codigo: str = "SPM"  # derivado do sufixo do projeto (-FD → FD, padrão SPM)
    titular_cpf_cnpj: Optional[str] = None
    categoria: Optional[str] = None
    projeto: Optional[str] = None
    valor_previsto: float = 0.0
    data_previsao: Optional[str] = None  # ISO date string
    observacao: Optional[str] = None
    linha_xlsx: int = 0  # para debug


class OrcamentoParser:
    """Parser do XLSX de orçamento com detecção de 6 seções."""

    SHEET_NAME = "Contas a Pagar"

    def detectar_secoes(
        self, arquivo: Union[str, Path, bytes]
    ) -> List[Tuple[int, NaturezaOrcamento]]:
        """
        Retorna lista de (linha_inicial, natureza) para cada seção detectada,
        em ordem de aparição no XLSX.
        """
        wb = self._abrir(arquivo)
        ws = wb[self.SHEET_NAME]

        secoes: List[Tuple[int, NaturezaOrcamento]] = []
        ja_detectadas = set()

        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, min(ws.max_column, 5) + 1):
                cell = ws.cell(row=row_idx, column=col_idx).value
                if not isinstance(cell, str):
                    continue
                for pattern, natureza in _SUBHEADERS:
                    if pattern.match(cell) and natureza not in ja_detectadas:
                        secoes.append((row_idx, natureza))
                        ja_detectadas.add(natureza)
                        break

        return secoes

    def _abrir(self, arquivo: Union[str, Path, bytes]):
        if isinstance(arquivo, (str, Path)):
            return openpyxl.load_workbook(arquivo, data_only=True)
        from io import BytesIO
        return openpyxl.load_workbook(BytesIO(arquivo), data_only=True)


# Colunas padrão por seção (índice 1-based do openpyxl).
# Todas as seções seguem o mesmo layout de colunas de 2 a 11, exceto PROFISSIONAIS
# que tem colunas: Previsão, CNPJ, Razão Social, Categoria, Projeto, RECEITA, DESPESAS, VALOR PAGO, Data, A PAGAR
_COL_DATA = 1
_COL_CNPJ = 2
_COL_RAZAO = 3
_COL_CATEGORIA = 4
_COL_PROJETO = 5
_COL_VALOR = 6            # "Valor da Conta" na maioria; "RECEITA" em PROFISSIONAIS
_COL_VALOR_LIQUIDO = 7    # "Valor Líquido"; "DESPESAS" em PROFISSIONAIS
_COL_VALOR_PAGO = 8
_COL_DATA_PAGAMENTO = 9
_COL_A_PAGAR = 10
_COL_OBS = 11


def _normalizar_data(valor) -> Optional[str]:
    if valor is None or valor == "":
        return None
    if hasattr(valor, "strftime"):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, str):
        # Ex: "Á CONFIRMAR", "À CONFIRMAR", datas em vários formatos
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})", valor.strip())
        if m:
            return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        return None
    return None


def _normalizar_cnpj(valor) -> Optional[str]:
    if valor is None:
        return None
    s = str(valor).strip()
    if not s:
        return None
    return s


def _linha_eh_valida(ws, row_idx: int, natureza: NaturezaOrcamento) -> bool:
    """Linha é válida se tem razão social e valor > 0."""
    razao = ws.cell(row=row_idx, column=_COL_RAZAO).value
    if natureza == NaturezaOrcamento.DESPESA_PROFISSIONAIS:
        valor = ws.cell(row=row_idx, column=_COL_VALOR_LIQUIDO).value  # DESPESAS
    else:
        valor = ws.cell(row=row_idx, column=_COL_VALOR_LIQUIDO).value  # Valor Líquido
        if valor in (None, 0):
            valor = ws.cell(row=row_idx, column=_COL_VALOR).value  # fallback Valor da Conta
    if not isinstance(razao, str) or not razao.strip():
        return False
    if not isinstance(valor, (int, float)) or valor <= 0:
        return False
    # Descarta cabeçalhos repetidos
    if razao.strip().lower() in {"razão social", "razao social", "cnpj/cpf"}:
        return False
    return True


def _extrair_uma_linha(ws, row_idx: int, natureza: NaturezaOrcamento) -> OrcamentoLinhaParsed:
    razao = ws.cell(row=row_idx, column=_COL_RAZAO).value
    cnpj = _normalizar_cnpj(ws.cell(row=row_idx, column=_COL_CNPJ).value)
    categoria = ws.cell(row=row_idx, column=_COL_CATEGORIA).value
    projeto = ws.cell(row=row_idx, column=_COL_PROJETO).value
    valor_liquido = ws.cell(row=row_idx, column=_COL_VALOR_LIQUIDO).value
    valor_conta = ws.cell(row=row_idx, column=_COL_VALOR).value
    valor = valor_liquido if isinstance(valor_liquido, (int, float)) and valor_liquido > 0 else valor_conta
    data_previsao = _normalizar_data(ws.cell(row=row_idx, column=_COL_DATA).value)
    obs = ws.cell(row=row_idx, column=_COL_OBS).value
    return OrcamentoLinhaParsed(
        natureza=natureza,
        titular_razao_social=str(razao).strip(),
        titular_cpf_cnpj=cnpj,
        categoria=str(categoria).strip() if isinstance(categoria, str) else None,
        projeto=str(projeto).strip() if isinstance(projeto, str) else None,
        valor_previsto=float(valor),
        data_previsao=data_previsao,
        observacao=str(obs).strip() if isinstance(obs, str) else None,
        linha_xlsx=row_idx,
    )


# Adicionar como método da classe OrcamentoParser:
def _extrair_linhas_secao_impl(self, arquivo, natureza: NaturezaOrcamento) -> List[OrcamentoLinhaParsed]:
    wb = self._abrir(arquivo)
    ws = wb[self.SHEET_NAME]
    secoes = self.detectar_secoes(arquivo)

    # Determinar intervalo de linhas dessa seção (do início até próxima seção ou EOF)
    secoes_ordenadas = sorted(secoes, key=lambda x: x[0])
    inicio = None
    fim = ws.max_row + 1
    for i, (row, nat) in enumerate(secoes_ordenadas):
        if nat == natureza:
            inicio = row + 1
            if i + 1 < len(secoes_ordenadas):
                fim = secoes_ordenadas[i + 1][0]
            break
    if inicio is None:
        return []

    # B5b: para DESPESA_PROFISSIONAIS (última seção nominal, mas XLSX real
    # do Hugo tem blocos extras após — INVESTIMENTOS, DESPESAS PESSOAIS DR HUGO,
    # Cartões de Crédito, NÃO IDENTIFICADO, linhas-resumo) — detectar fim via
    # sentinelas textuais e/ou linhas vazias consecutivas.
    linhas: List[OrcamentoLinhaParsed] = []
    linhas_vazias_consecutivas = 0
    for row_idx in range(inicio, fim):
        if natureza == NaturezaOrcamento.DESPESA_PROFISSIONAIS:
            if _eh_fim_secao_profissionais(ws, row_idx):
                break
            # Contar linhas completamente vazias (B-K) para detectar fim por separador
            vazia = all(
                ws.cell(row=row_idx, column=c).value in (None, "")
                for c in range(2, 12)
            )
            if vazia:
                linhas_vazias_consecutivas += 1
                if linhas_vazias_consecutivas >= 3:
                    break
                continue
            else:
                linhas_vazias_consecutivas = 0
        if _linha_eh_valida(ws, row_idx, natureza):
            linhas.append(_extrair_uma_linha(ws, row_idx, natureza))
    return linhas


OrcamentoParser.extrair_linhas_secao = _extrair_linhas_secao_impl


@dataclass
class ResultadoParse:
    linhas: List[OrcamentoLinhaParsed]
    total_linhas: int
    linhas_por_secao: dict[NaturezaOrcamento, int]
    linhas_descartadas: int
    avisos: List[str]


def _derivar_empresa(projeto: Optional[str]) -> str:
    """Deriva empresa do sufixo do projeto: '-FD' → FD, padrão SPM."""
    if not projeto:
        return "SPM"
    p = projeto.upper().strip()
    if p.endswith("-FD") or p.endswith(" FD") or "- FD" in p or "-FD " in p:
        return "FD"
    return "SPM"


def _parse_completo_impl(self, arquivo) -> ResultadoParse:
    todas_linhas: List[OrcamentoLinhaParsed] = []
    por_secao: dict[NaturezaOrcamento, int] = {}
    avisos: List[str] = []
    descartadas = 0

    for natureza in NaturezaOrcamento:
        linhas = self.extrair_linhas_secao(arquivo, natureza)
        for l in linhas:
            l.empresa_codigo = _derivar_empresa(l.projeto)
            todas_linhas.append(l)
        por_secao[natureza] = len(linhas)

    return ResultadoParse(
        linhas=todas_linhas,
        total_linhas=len(todas_linhas),
        linhas_por_secao=por_secao,
        linhas_descartadas=descartadas,
        avisos=avisos,
    )


OrcamentoParser.parse_completo = _parse_completo_impl
