"""
Parser do Relatório Financeiro Bonificado do Pega Plantao.

Estrutura do XLSX:
  - Linhas 1-3: cabeçalho global (ignorar)
  - Linha vazia separa blocos de profissionais
  - Cada bloco:
    - Linha A: "NOME PRESTADOR  -  CRM/UF"
    - Linha B: "Transação: PIX  Tipo de Documento: CPF/CNPJ  Chave Pix: XXX  Documento: XXX  Razão social: NOME"
    - Linha C: cabeçalho colunas (Data, Local, Tipo, Duração (h), Valor, Total, Total Pago, Saldo)
    - Linhas D+: plantões (Data | Local | Tipo | Duração | Valor | fórmula | Total Pago | Saldo)
    - Linha "Total": subtotal — Saldo desta linha = saldo_final do bloco

REGRAS CRÍTICAS:
  - Coluna contrato = "Local" (NUNCA "Setor")
  - Saldo já tem taxas e retenções deduzidas — NUNCA recalcular
"""

from __future__ import annotations

import re
import logging
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Optional, Union

import openpyxl

logger = logging.getLogger(__name__)

# Índices das colunas (0-based): Data, Local, Tipo, Duração (h), Valor, Total, Total Pago, Saldo
COL_DATA = 0
COL_LOCAL = 1   # Contrato — NUNCA usar "Setor"
COL_TIPO = 2
COL_DURACAO = 3
COL_VALOR = 4
COL_TOTAL = 5
COL_TOTAL_PAGO = 6
COL_SALDO = 7


def extrair_cabecalho_prestador(linha_a: str) -> Dict[str, str]:
    """
    Extrai nome, CRM e UF da linha A do bloco.
    Formato: "NOME PRESTADOR  -  CRM/UF"
    Ex: "JOAO DA SILVA  -  123456/SP"
    """
    linha_a = (linha_a or "").strip()
    partes = re.split(r'\s+-\s+', linha_a, maxsplit=1)
    nome = partes[0].strip()
    crm, uf = "", ""
    if len(partes) == 2:
        crm_uf = partes[1].strip()
        m = re.match(r'([\d]+)\s*/\s*([A-Z]{2})', crm_uf)
        if m:
            crm, uf = m.group(1), m.group(2)
        else:
            m2 = re.match(r'(?:CRM\s*)?([\d]+)\s+([A-Z]{2})', crm_uf)
            if m2:
                crm, uf = m2.group(1), m2.group(2)
            else:
                crm = crm_uf
    return {"nome_prestador": nome, "crm": crm, "uf": uf}


def extrair_pix_info(linha_b: str) -> Dict[str, str]:
    """
    Extrai dados PIX da linha B do bloco.
    Formato: "Transação: PIX  Tipo de Documento: CPF  Chave Pix: XXX  Documento: XXX  Razão social: ..."
    """
    linha_b = (linha_b or "").strip()
    result = {"tipo_doc": "", "chave_pix": "", "documento": "", "razao_social_pj": ""}

    m = re.search(r'Tipo\s+de\s+Documento:\s*(CPF|CNPJ)', linha_b, re.IGNORECASE)
    if m:
        result["tipo_doc"] = m.group(1).upper()

    # Chave Pix: até dois espaços ou outro campo
    m = re.search(r'Chave\s+Pix:\s*(.+?)(?=\s{2,}|\s*Documento:|\s*Raz)', linha_b, re.IGNORECASE)
    if m:
        result["chave_pix"] = m.group(1).strip()

    # Documento: próximo token sem espaços
    m = re.search(r'\bDocumento:\s*([^\s]+)', linha_b, re.IGNORECASE)
    if m:
        result["documento"] = m.group(1).strip()

    # Razão social: resto da linha após "Razão social:"
    m = re.search(r'Raz[aã]o\s+social:\s*(.*?)\s*$', linha_b, re.IGNORECASE)
    if m:
        result["razao_social_pj"] = m.group(1).strip()

    return result


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".").replace(" ", "").replace("R$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_mes_competencia(value: Any) -> str:
    if value is None:
        return ""
    import datetime
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%m/%Y")
    s = str(value).strip()
    m = re.match(r'(\d{2})/(\d{2})/(\d{4})', s)
    if m:
        return f"{m.group(2)}/{m.group(3)}"
    m2 = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m2:
        return f"{m2.group(2)}/{m2.group(1)}"
    return s


def _row_values(row) -> list:
    return [cell.value for cell in row]


def _is_empty_row(values: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def _is_header_cols_row(values: list) -> bool:
    """Linha C: cabeçalho 'Data, Local, Tipo...'"""
    if not values or values[0] is None:
        return False
    return str(values[0]).strip().lower() == "data"


def _is_total_row(values: list) -> bool:
    if not values or values[0] is None:
        return False
    return str(values[0]).strip().lower() == "total"


def _is_prestador_header(values: list) -> bool:
    """Linha A: só primeira célula preenchida, texto com nome do prestador."""
    non_none = [v for v in values if v is not None and str(v).strip()]
    if len(non_none) != 1:
        return False
    s = str(non_none[0]).strip()
    # Deve ter " - " (separador CRM) ou ser texto longo em maiúsculas
    if not s or len(s) < 3:
        return False
    # Não é data
    if re.match(r'\d{2}/\d{2}/\d{4}', s):
        return False
    # Não é cabeçalho de colunas
    if s.lower() in ("data", "total", "local", "tipo"):
        return False
    # Tem traço separando nome do CRM, ou é texto que parece nome
    return " - " in s or (s.isupper() and len(s) > 5)


def _is_pix_info_row(values: list) -> bool:
    """Linha B: contém 'Transação' ou 'Tipo de Documento'."""
    if not values or values[0] is None:
        return False
    s = str(values[0]).strip()
    return bool(re.search(r'transa[cç][aã]o|tipo.*documento|chave.*pix', s, re.IGNORECASE))


def parse_relatorio(source: Union[str, Path, bytes, BytesIO]) -> List[Dict[str, Any]]:
    """
    Parseia o Relatório Financeiro Bonificado do Pega Plantao.

    Aceita path (str/Path), bytes ou BytesIO — upload via HTTP usa bytes/BytesIO.

    Retorna lista de dicts, um por prestador+contrato:
      nome_prestador, crm, uf, contrato, mes_competencia,
      saldo, tipo_doc, razao_social_pj, chave_pix, documento
    """
    if isinstance(source, (str, Path)):
        filepath = Path(source)
        if not filepath.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")
        logger.info(f"Abrindo PP XLSX: {filepath}")
        wb = openpyxl.load_workbook(filepath, data_only=True)
    else:
        if isinstance(source, bytes):
            source = BytesIO(source)
        logger.info("Abrindo PP XLSX: <bytes/BytesIO>")
        wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows())

    results = []
    SKIP_ROWS = 3  # Ignora linhas 1-3 (cabeçalho global)
    i = SKIP_ROWS

    while i < len(rows):
        row_vals = _row_values(rows[i])

        if _is_empty_row(row_vals):
            i += 1
            continue

        if not _is_prestador_header(row_vals):
            i += 1
            continue

        # === Linha A: nome do prestador ===
        linha_a = str(row_vals[0]).strip()
        cabecalho = extrair_cabecalho_prestador(linha_a)
        logger.debug(f"Bloco prestador: {cabecalho['nome_prestador']}")
        i += 1

        # === Linha B: info PIX ===
        pix_info = {"tipo_doc": "", "chave_pix": "", "documento": "", "razao_social_pj": ""}
        while i < len(rows):
            v = _row_values(rows[i])
            if _is_empty_row(v):
                i += 1
                continue
            if _is_pix_info_row(v):
                pix_info = extrair_pix_info(str(v[0]).strip())
                i += 1
            break

        # === Linha C: cabeçalho de colunas (pular) ===
        while i < len(rows):
            v = _row_values(rows[i])
            if _is_empty_row(v):
                i += 1
                continue
            if _is_header_cols_row(v):
                i += 1
            break

        # === Linhas D+: dados de plantão ===
        # REGRA (CLAUDE.md regra 1): 1 arquivo PP = 1 contrato = 1 remessa por prestador.
        # A coluna "Local" indica unidade/setor dentro do MESMO contrato — NUNCA fragmentar.
        # Emitir sempre 1 registro por prestador com saldo = linha Total.
        locais_visitados: list = []          # locais distintos na ordem de aparição
        saldo_acumulado_total: float = 0.0   # fallback se Total não tiver cache
        mes_bloco: str = ""                  # primeiro mês detectado no bloco

        while i < len(rows):
            v = _row_values(rows[i])

            if _is_empty_row(v):
                i += 1
                break  # Fim do bloco deste prestador

            if _is_total_row(v):
                saldo_total = _to_float(v[COL_SALDO]) if len(v) > COL_SALDO else None
                # Fórmula Excel sem cache retorna None — usar soma acumulada como fallback
                saldo_final = saldo_total if (saldo_total is not None and saldo_total != 0.0) \
                    else saldo_acumulado_total

                # Constrói string do contrato preservando prefixo "UF - CIDADE - "
                # e concatenando apenas os sub-locais quando há múltiplos locais distintos.
                locais_unicos = list(dict.fromkeys(locais_visitados))
                if not locais_unicos:
                    contrato_str = ""
                elif len(locais_unicos) == 1:
                    contrato_str = locais_unicos[0]
                else:
                    # Tenta extrair prefixo "UF - CIDADE" comum a todos os locais
                    # para produzir "UF - CIDADE - SUB_A + SUB_B" (compatível com _parse_local_pp)
                    parts_primeiro = [p.strip() for p in locais_unicos[0].replace('–', '-').split(' - ')]
                    if len(parts_primeiro) >= 2:
                        prefixo = f"{parts_primeiro[0]} - {parts_primeiro[1]}"
                        # Verifica se todos os locais compartilham o mesmo prefixo UF/CIDADE
                        todos_mesma_cidade = all(
                            loc.replace('–', '-').startswith(prefixo)
                            for loc in locais_unicos
                        )
                        if todos_mesma_cidade:
                            # Extrai sub-local de cada um (tudo após "UF - CIDADE - ")
                            sub_locais = []
                            prefixo_com_sep = prefixo + ' - '
                            for loc in locais_unicos:
                                loc_norm = loc.replace('–', '-')
                                if loc_norm.startswith(prefixo_com_sep):
                                    sub_locais.append(loc_norm[len(prefixo_com_sep):])
                                else:
                                    sub_locais.append(loc_norm[len(prefixo):].lstrip(' -'))
                            if len(sub_locais) <= 3:
                                contrato_str = prefixo + ' - ' + ' + '.join(sub_locais)
                            else:
                                contrato_str = prefixo + ' - ' + sub_locais[0] + f' + {len(sub_locais)-1} outros'
                        else:
                            # UF/cidades diferentes (não deveria ocorrer, mas defensivo)
                            contrato_str = locais_unicos[0]
                    else:
                        # Formato inesperado — concatena inteiros
                        if len(locais_unicos) <= 3:
                            contrato_str = ' + '.join(locais_unicos)
                        else:
                            contrato_str = locais_unicos[0] + f' + {len(locais_unicos)-1} outros'

                record = {
                    "nome_prestador": cabecalho["nome_prestador"],
                    "crm": cabecalho["crm"],
                    "uf": cabecalho["uf"],
                    "contrato": contrato_str,
                    "mes_competencia": mes_bloco,
                    "saldo": saldo_final,
                    "tipo_doc": pix_info["tipo_doc"],
                    "razao_social_pj": pix_info["razao_social_pj"],
                    "chave_pix": pix_info["chave_pix"],
                    "documento": pix_info["documento"],
                }
                results.append(record)
                logger.debug(
                    f"  Registro: {record['nome_prestador']} | {contrato_str} "
                    f"| locais={len(locais_unicos)} | saldo={saldo_final:.2f}"
                )

                # Reset para próximo bloco (pode haver novo cabeçalho de colunas)
                locais_visitados = []
                saldo_acumulado_total = 0.0
                mes_bloco = ""
                i += 1

                # Após Total, pode vir novo cabeçalho de colunas ou linha vazia
                continue

            if _is_header_cols_row(v):
                # Cabeçalho de colunas dentro do bloco (ignorar)
                i += 1
                continue

            # Linha de plantão ou taxa: registra local e acumula saldo
            if len(v) > COL_LOCAL and v[COL_DATA] is not None:
                local_str = str(v[COL_LOCAL]).strip() if v[COL_LOCAL] else ""
                if local_str and local_str not in locais_visitados:
                    locais_visitados.append(local_str)
                # Captura mês da primeira data válida do bloco
                if not mes_bloco:
                    mes_bloco = _to_mes_competencia(v[COL_DATA])

            # Acumula coluna H (Saldo) de QUALQUER linha numérica —
            # cobre plantões (positivos) e taxas SPM (negativas, sem data em col 0).
            # Fallback usado quando a fórmula =SUM da linha Total vem sem cache.
            if len(v) > COL_SALDO:
                val_h = _to_float(v[COL_SALDO])
                if val_h is not None:
                    saldo_acumulado_total += val_h

            i += 1

    logger.info(f"Parse PP concluído: {len(results)} registros")
    return results
