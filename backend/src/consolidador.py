# spm-faturas: Consolidador multi-cartão

from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from src.excel_builder import (
    CORES, FILL_HEADER, FILL_VERDE, FILL_VERMELHO, FILL_AMARELO, FILL_AZUL,
    FILL_CINZA, FILL_PREVIA, FILL_SUBTOTAL_PJ, FILL_SUBTOTAL_PF,
    FILL_CINZA_ESTORNO, FILL_ALERTA,
    FONT_HEADER, FONT_BOLD, FONT_NORMAL, FONT_SUBTOTAL_PJ, FONT_SUBTOTAL_PF,
    THIN_BORDER, ALIGN_CENTER, ALIGN_LEFT, FMT_BRL, TIPOS_PF_PJ,
    _escrever_header, _escrever_header_em_row, _aplicar_estilos,
    _calcular_ultima_fatura, _escrever_secao_categoria,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# API publica
# ---------------------------------------------------------------------------

def gerar_consolidado(
    faturas_processadas: list[dict],
    config: dict,
    caminho_saida: str,
) -> str:
    """Entry point. Gera Excel consolidado multi-cartao.

    faturas_processadas: lista de dicts com:
        fatura: retorno do extrator
        classificacao: retorno do classificador
    Retorna caminho do arquivo gerado.
    """
    wb = Workbook()

    mes_referencia = ""
    for fp in faturas_processadas:
        mr = fp["fatura"].get("resumo", {}).get("mes_referencia", "")
        if mr:
            mes_referencia = mr
            break

    _aba_painel_executivo(wb, faturas_processadas, mes_referencia)
    _aba_projecao_consolidada(wb, faturas_processadas, mes_referencia, config)
    _aba_pj_consolidado(wb, faturas_processadas)
    _aba_relatorio_conselho_consolidado(wb, faturas_processadas, mes_referencia, config)

    # Remover sheet default
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    path = Path(caminho_saida)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info("Consolidado salvo em %s", caminho_saida)
    return str(path)


def montar_caminho_consolidado(config: dict, mes_referencia: str) -> str:
    """Gera caminho do consolidado:
    {output_base}/{ano}/{mes}.{ano}/Hugo - Consolidado - Cartoes - {Mes}.{YYYY}.xlsx
    """
    output_base = config.get("caminhos", {}).get("output_base", "./output")
    output_base = str(Path(output_base).expanduser())
    titular = config.get("titular", "Hugo")
    meses_nomes = config.get("meses", {})

    if mes_referencia and len(mes_referencia) >= 7:
        ano = mes_referencia[:4]
        mes_num = int(mes_referencia[5:7])
    else:
        hoje = datetime.now()
        ano = str(hoje.year)
        mes_num = hoje.month

    nome_mes = meses_nomes.get(mes_num, f"Mes{mes_num:02d}")
    pasta = Path(output_base) / ano / f"{nome_mes}.{ano}"
    nome_arquivo = f"{titular} - Consolidado - Cartoes - {nome_mes}.{ano}.xlsx"
    return str(pasta / nome_arquivo)


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _todos_lancamentos(fp: dict) -> list[dict]:
    """Retorna todos os lancamentos (classificados + pendentes) de uma fatura processada."""
    classif = fp.get("classificacao", {})
    return classif.get("classificados", []) + classif.get("pendentes", [])


def _todos_lancamentos_multi(faturas: list[dict]) -> list[dict]:
    """Retorna todos os lancamentos de todas as faturas."""
    result = []
    for fp in faturas:
        result.extend(_todos_lancamentos(fp))
    return result


# ---------------------------------------------------------------------------
# Abas
# ---------------------------------------------------------------------------

def _aba_painel_executivo(
    wb: Workbook, faturas: list[dict], mes_referencia: str
) -> None:
    """Aba 'Painel Executivo' — resumo de todos os cartoes."""
    ws = wb.create_sheet(title="Painel Executivo")

    headers = [
        "Banco", "Cartão", "Final", "Vencimento", "Total Fatura (R$)",
        "Total PF (R$)", "Total PJ (R$)", "Encargos (R$)", "Pendentes",
    ]
    _escrever_header(ws, headers)

    totais = {
        "total_fatura": 0.0,
        "total_pf": 0.0,
        "total_pj": 0.0,
        "encargos": 0.0,
        "pendentes": 0,
    }

    row = 2
    for fp in faturas:
        fatura = fp["fatura"]
        classif = fp["classificacao"]
        resumo = fatura.get("resumo", {})
        stats = classif.get("stats", {})

        banco = fatura.get("banco", "")
        cartao = fatura.get("cartao", "")
        final = fatura.get("final", "")
        vencimento = resumo.get("vencimento", "")
        total_fatura = resumo.get("total_fatura", 0.0)
        encargos = resumo.get("encargos", 0.0)
        pend = stats.get("pendentes", 0)

        todos = _todos_lancamentos(fp)
        total_pf = sum(l.get("valor", 0.0) for l in todos if l.get("classificacao") == "PF" and l.get("tipo") in TIPOS_PF_PJ)
        total_pj = sum(l.get("valor", 0.0) for l in todos if l.get("classificacao") == "PJ" and l.get("tipo") in TIPOS_PF_PJ)

        ws.cell(row=row, column=1, value=banco)
        ws.cell(row=row, column=2, value=cartao)
        ws.cell(row=row, column=3, value=final)
        ws.cell(row=row, column=4, value=vencimento)
        ws.cell(row=row, column=5, value=total_fatura).number_format = FMT_BRL
        ws.cell(row=row, column=6, value=total_pf).number_format = FMT_BRL
        ws.cell(row=row, column=7, value=total_pj).number_format = FMT_BRL
        ws.cell(row=row, column=8, value=encargos).number_format = FMT_BRL
        ws.cell(row=row, column=9, value=pend)

        if pend > 0:
            ws.cell(row=row, column=9).fill = FILL_AMARELO

        if row % 2 == 0:
            for col in range(1, 10):
                if ws.cell(row=row, column=col).fill != FILL_AMARELO:
                    ws.cell(row=row, column=col).fill = FILL_CINZA

        totais["total_fatura"] += total_fatura
        totais["total_pf"] += total_pf
        totais["total_pj"] += total_pj
        totais["encargos"] += encargos
        totais["pendentes"] += pend

        row += 1

    # Linha de totais consolidados
    ws.cell(row=row, column=1, value="TOTAL CONSOLIDADO").font = FONT_BOLD
    ws.cell(row=row, column=4, value="").font = FONT_BOLD
    ws.cell(row=row, column=5, value=totais["total_fatura"]).number_format = FMT_BRL
    ws.cell(row=row, column=5).font = FONT_BOLD
    ws.cell(row=row, column=6, value=totais["total_pf"]).number_format = FMT_BRL
    ws.cell(row=row, column=6).font = FONT_BOLD
    ws.cell(row=row, column=7, value=totais["total_pj"]).number_format = FMT_BRL
    ws.cell(row=row, column=7).font = FONT_BOLD
    ws.cell(row=row, column=8, value=totais["encargos"]).number_format = FMT_BRL
    ws.cell(row=row, column=8).font = FONT_BOLD
    ws.cell(row=row, column=9, value=totais["pendentes"]).font = FONT_BOLD

    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = FILL_PREVIA

    _aplicar_estilos(ws, num_colunas=9, num_linhas=row)
    ws.sheet_properties.tabColor = CORES["header"]


def _aba_projecao_consolidada(
    wb: Workbook,
    faturas: list[dict],
    mes_referencia: str,
    config: dict,
) -> None:
    """Aba 'Projecao Consolidada' — parcelas futuras de todos os cartoes + recorrencias."""
    ws = wb.create_sheet(title="Projecao Consolidada")

    if not mes_referencia:
        ws.cell(row=1, column=1, value="Sem mes de referencia.")
        return

    # Coletar parceladas de todas as faturas
    parcelas_por_grupo = defaultdict(lambda: {"valor": 0.0, "max_rest": 0, "banco": ""})

    for fp in faturas:
        banco = fp["fatura"].get("banco", "")
        todos = _todos_lancamentos(fp)
        for lanc in todos:
            if lanc.get("tipo") != "parcelada":
                continue
            rest = lanc.get("parcelas_restantes") or 0
            if rest <= 0:
                continue
            desc = lanc.get("descricao", "")
            chave = f"{banco} | {desc}"
            parcelas_por_grupo[chave]["valor"] += lanc.get("valor", 0.0)
            parcelas_por_grupo[chave]["banco"] = banco
            if rest > parcelas_por_grupo[chave]["max_rest"]:
                parcelas_por_grupo[chave]["max_rest"] = rest

    # Adicionar recorrencias do config
    recorrencias = config.get("recorrencias", [])
    ano_ref = int(mes_referencia[:4])
    mes_ref = int(mes_referencia[5:7])

    for rec in recorrencias:
        nome = rec.get("nome", "Recorrencia")
        cartao = rec.get("cartao", "")
        valor_atual = rec.get("valor_atual", 0.0)
        valor_desc = rec.get("valor_com_desconto")
        desconto_a_partir = rec.get("desconto_a_partir", "")
        encerramento = rec.get("encerramento", "")

        # Calcular meses restantes ate encerramento
        if encerramento and len(encerramento) >= 7:
            enc_ano = int(encerramento[:4])
            enc_mes = int(encerramento[5:7])
            meses_rest = (enc_ano - ano_ref) * 12 + (enc_mes - mes_ref)
            if meses_rest <= 0:
                continue
        else:
            meses_rest = 12  # default 1 ano

        chave = f"{cartao} | {nome} (recorrencia)"
        parcelas_por_grupo[chave] = {
            "valor": valor_atual,
            "valor_desc": valor_desc,
            "desconto_a_partir": desconto_a_partir,
            "banco": cartao,
            "max_rest": meses_rest,
            "recorrencia": True,
        }

    if not parcelas_por_grupo:
        ws.cell(row=1, column=1, value="Sem parcelamentos ou recorrencias futuras.")
        return

    # Calcular max meses
    max_meses = max(g["max_rest"] for g in parcelas_por_grupo.values())
    max_meses = min(max_meses, 24)  # limitar a 24 meses

    # Gerar lista de meses futuros
    meses_futuros = []
    ano, mes = ano_ref, mes_ref
    for _ in range(max_meses):
        mes += 1
        if mes > 12:
            mes = 1
            ano += 1
        meses_futuros.append(f"{ano:04d}-{mes:02d}")

    headers = ["Cartão | Descrição", "Valor Mensal (R$)"] + meses_futuros + ["Total Futuro (R$)"]
    _escrever_header(ws, headers)

    row = 2
    total_por_mes = [0.0] * len(meses_futuros)
    total_geral = 0.0

    for chave in sorted(parcelas_por_grupo.keys()):
        grupo = parcelas_por_grupo[chave]
        is_rec = grupo.get("recorrencia", False)
        valor_base = grupo["valor"]
        max_rest = grupo["max_rest"]

        ws.cell(row=row, column=1, value=chave)
        ws.cell(row=row, column=2, value=valor_base).number_format = FMT_BRL

        total_linha = 0.0
        for j, mes_futuro in enumerate(meses_futuros):
            if j < max_rest:
                # Recorrencias podem ter desconto a partir de uma data
                valor_mes = valor_base
                if is_rec and grupo.get("valor_desc") and grupo.get("desconto_a_partir"):
                    if mes_futuro >= grupo["desconto_a_partir"]:
                        valor_mes = grupo["valor_desc"]

                ws.cell(row=row, column=3 + j, value=valor_mes).number_format = FMT_BRL
                total_por_mes[j] += valor_mes
                total_linha += valor_mes
            else:
                ws.cell(row=row, column=3 + j, value="")

        col_total = 3 + len(meses_futuros)
        ws.cell(row=row, column=col_total, value=total_linha).number_format = FMT_BRL
        total_geral += total_linha

        # Destaque recorrencias
        if is_rec:
            for col in range(1, col_total + 1):
                ws.cell(row=row, column=col).fill = FILL_AMARELO
        elif row % 2 == 0:
            for col in range(1, col_total + 1):
                ws.cell(row=row, column=col).fill = FILL_CINZA

        row += 1

    # Linha de totais
    ws.cell(row=row, column=1, value="TOTAL").font = FONT_BOLD
    for j, total_mes in enumerate(total_por_mes):
        cell = ws.cell(row=row, column=3 + j, value=total_mes)
        cell.number_format = FMT_BRL
        cell.font = FONT_BOLD
    col_total = 3 + len(meses_futuros)
    cell = ws.cell(row=row, column=col_total, value=total_geral)
    cell.number_format = FMT_BRL
    cell.font = FONT_BOLD

    for col in range(1, col_total + 1):
        ws.cell(row=row, column=col).fill = FILL_PREVIA

    _aplicar_estilos(ws, num_colunas=col_total, num_linhas=row)
    ws.sheet_properties.tabColor = CORES["previa"]


def _aba_pj_consolidado(wb: Workbook, faturas: list[dict]) -> None:
    """Aba 'PJ Consolidado' — despesas PJ de todos os cartoes por categoria."""
    ws = wb.create_sheet(title="PJ Consolidado")

    headers = [
        "Banco", "Cartão", "Data", "Descrição", "Valor (R$)", "Categoria",
    ]
    _escrever_header(ws, headers)

    row = 2
    totais_cat = defaultdict(float)
    total_geral = 0.0

    for fp in faturas:
        banco = fp["fatura"].get("banco", "")
        cartao = fp["fatura"].get("cartao", "")
        todos = _todos_lancamentos(fp)

        pj_lancs = [l for l in todos if l.get("classificacao") == "PJ" and l.get("tipo") in TIPOS_PF_PJ]
        pj_lancs.sort(key=lambda l: l.get("data", ""))

        for lanc in pj_lancs:
            cat = lanc.get("categoria", "Outros")
            valor = lanc.get("valor", 0.0)

            ws.cell(row=row, column=1, value=banco)
            ws.cell(row=row, column=2, value=cartao)
            ws.cell(row=row, column=3, value=lanc.get("data", ""))
            ws.cell(row=row, column=4, value=lanc.get("descricao", ""))
            ws.cell(row=row, column=5, value=valor).number_format = FMT_BRL
            ws.cell(row=row, column=6, value=cat)

            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = FILL_VERDE

            totais_cat[cat] += valor
            total_geral += valor
            row += 1

    # Resumo por categoria
    row += 1
    ws.cell(row=row, column=1, value="Resumo por Categoria").font = Font(bold=True, size=12)
    row += 1
    _escrever_header(ws, ["Categoria", "Total (R$)", "% do PJ"], row=row)
    row += 1

    for cat in sorted(totais_cat.keys()):
        val = totais_cat[cat]
        pct = (val / total_geral * 100) if total_geral else 0.0
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=val).number_format = FMT_BRL
        ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        if row % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = FILL_CINZA
        row += 1

    # Total geral
    ws.cell(row=row, column=1, value="TOTAL PJ").font = FONT_BOLD
    ws.cell(row=row, column=2, value=total_geral).number_format = FMT_BRL
    ws.cell(row=row, column=2).font = FONT_BOLD
    ws.cell(row=row, column=3, value="100.0%").font = FONT_BOLD
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = FILL_SUBTOTAL_PJ
        ws.cell(row=row, column=col).font = FONT_SUBTOTAL_PJ

    _aplicar_estilos(ws, num_colunas=6, num_linhas=row)
    ws.sheet_properties.tabColor = CORES["verde_pj"]


def _aba_relatorio_conselho_consolidado(
    wb: Workbook,
    faturas: list[dict],
    mes_referencia: str,
    config: dict,
) -> None:
    """Aba 'Relatorio Conselho' — documento formal consolidado."""
    ws = wb.create_sheet(title="Relatorio Conselho")

    # Titulo
    ws.merge_cells("A1:E1")
    cell = ws.cell(
        row=1, column=1,
        value="Relatório Consolidado de Prestação de Contas — Cartões Corporativos",
    )
    cell.font = Font(bold=True, size=14, color=CORES["header"])
    cell.alignment = ALIGN_CENTER

    ws.merge_cells("A2:E2")
    ws.cell(row=2, column=1, value=f"Referência: {mes_referencia}").alignment = ALIGN_CENTER

    # Resumo por cartao
    row = 4
    ws.cell(row=row, column=1, value="Resumo por Cartão").font = Font(bold=True, size=12)
    row += 1

    _escrever_header(ws, ["Banco", "Total Fatura", "PF", "PJ", "Encargos"], row=row)
    row += 1

    total_fatura_geral = 0.0
    total_pf_geral = 0.0
    total_pj_geral = 0.0
    total_enc_geral = 0.0

    for fp in faturas:
        fatura = fp["fatura"]
        resumo = fatura.get("resumo", {})
        banco = fatura.get("banco", "")
        tf = resumo.get("total_fatura", 0.0)
        enc = resumo.get("encargos", 0.0)

        todos = _todos_lancamentos(fp)
        pf = sum(l.get("valor", 0.0) for l in todos if l.get("classificacao") == "PF" and l.get("tipo") in TIPOS_PF_PJ)
        pj = sum(l.get("valor", 0.0) for l in todos if l.get("classificacao") == "PJ" and l.get("tipo") in TIPOS_PF_PJ)

        ws.cell(row=row, column=1, value=banco).font = FONT_BOLD
        ws.cell(row=row, column=2, value=tf).number_format = FMT_BRL
        ws.cell(row=row, column=3, value=pf).number_format = FMT_BRL
        ws.cell(row=row, column=4, value=pj).number_format = FMT_BRL
        ws.cell(row=row, column=5, value=enc).number_format = FMT_BRL

        total_fatura_geral += tf
        total_pf_geral += pf
        total_pj_geral += pj
        total_enc_geral += enc
        row += 1

    # Totais consolidados
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = FILL_PREVIA
    ws.cell(row=row, column=1, value="CONSOLIDADO").font = FONT_BOLD
    ws.cell(row=row, column=2, value=total_fatura_geral).number_format = FMT_BRL
    ws.cell(row=row, column=2).font = FONT_BOLD
    ws.cell(row=row, column=3, value=total_pf_geral).number_format = FMT_BRL
    ws.cell(row=row, column=3).font = FONT_BOLD
    ws.cell(row=row, column=4, value=total_pj_geral).number_format = FMT_BRL
    ws.cell(row=row, column=4).font = FONT_BOLD
    ws.cell(row=row, column=5, value=total_enc_geral).number_format = FMT_BRL
    ws.cell(row=row, column=5).font = FONT_BOLD
    row += 2

    # Distribuicao PF vs PJ
    ws.cell(row=row, column=1, value="Distribuição PF vs PJ").font = Font(bold=True, size=12)
    row += 1

    total_class = total_pf_geral + total_pj_geral
    pct_pf = (total_pf_geral / total_class * 100) if total_class else 0.0
    pct_pj = (total_pj_geral / total_class * 100) if total_class else 0.0

    ws.cell(row=row, column=1, value="Pessoa Física (PF)").font = FONT_BOLD
    ws.cell(row=row, column=2, value=total_pf_geral).number_format = FMT_BRL
    ws.cell(row=row, column=3, value=f"{pct_pf:.1f}%")
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = FILL_AZUL
    row += 1

    ws.cell(row=row, column=1, value="Pessoa Jurídica (PJ)").font = FONT_BOLD
    ws.cell(row=row, column=2, value=total_pj_geral).number_format = FMT_BRL
    ws.cell(row=row, column=3, value=f"{pct_pj:.1f}%")
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = FILL_VERDE
    row += 2

    # Detalhamento PJ por categoria (consolidado)
    ws.cell(row=row, column=1, value="Detalhamento PJ por Categoria (Consolidado)").font = Font(bold=True, size=12)
    row += 1

    todos_lancs = _todos_lancamentos_multi(faturas)
    cats_pj = defaultdict(float)
    for lanc in todos_lancs:
        if lanc.get("classificacao") == "PJ" and lanc.get("tipo") in TIPOS_PF_PJ:
            cats_pj[lanc.get("categoria", "Outros")] += lanc.get("valor", 0.0)

    _escrever_header(ws, ["Categoria", "Valor (R$)", "% PJ"], row=row)
    row += 1

    for cat in sorted(cats_pj.keys()):
        val = cats_pj[cat]
        pct = (val / total_pj_geral * 100) if total_pj_geral else 0.0
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=val).number_format = FMT_BRL
        ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        if row % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = FILL_CINZA
        row += 1

    # Recorrencias
    recorrencias = config.get("recorrencias", [])
    if recorrencias:
        row += 1
        ws.cell(row=row, column=1, value="Recorrências Ativas").font = Font(bold=True, size=12)
        row += 1
        _escrever_header(ws, ["Nome", "Cartão", "Valor Mensal (R$)", "Encerramento"], row=row)
        row += 1
        for rec in recorrencias:
            ws.cell(row=row, column=1, value=rec.get("nome", ""))
            ws.cell(row=row, column=2, value=rec.get("cartao", ""))
            ws.cell(row=row, column=3, value=rec.get("valor_atual", 0.0)).number_format = FMT_BRL
            ws.cell(row=row, column=4, value=rec.get("encerramento", ""))
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = FILL_AMARELO
            row += 1

    _aplicar_estilos(ws, num_colunas=5, num_linhas=row - 1)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.sheet_properties.tabColor = CORES["verde_pj"]


# ===========================================================================
# CONSOLIDADO DE EXTRATOS BANCARIOS
# ===========================================================================

def gerar_consolidado_extratos(
    extratos_processados: list[dict],
    config: dict,
    caminho_saida: str,
) -> str:
    """Entry point. Gera Excel consolidado de extratos bancarios.

    extratos_processados: lista de dicts com:
        extrato: retorno do extrator (extrair_extrato_bb/itau/unicred/sisprime)
        classificacao: retorno do classificador
    Retorna caminho do arquivo gerado.
    """
    wb = Workbook()

    mes_referencia = ""
    for ep in extratos_processados:
        mr = ep["extrato"].get("mes_referencia", "")
        if mr:
            mes_referencia = mr
            break

    _aba_painel_contas(wb, extratos_processados)
    _aba_fluxo_consolidado(wb, extratos_processados)
    _aba_receitas_despesas(wb, extratos_processados)
    _aba_relatorio_executivo(wb, extratos_processados, mes_referencia)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # Montar caminho do arquivo
    meses_config = config.get("meses", {})
    titular = config.get("titular", "Hugo")

    if mes_referencia and len(mes_referencia) >= 7:
        ano = mes_referencia[:4]
        mes_num = int(mes_referencia[5:7])
        mes_nome = meses_config.get(mes_num, f"{mes_num:02d}")
        nome_arquivo = f"{titular} - Extratos - Consolidado - {mes_nome}.{ano}.xlsx"
        subpasta = f"{ano}/{mes_nome}.{ano}"
    else:
        nome_arquivo = f"{titular} - Extratos - Consolidado.xlsx"
        subpasta = ""

    caminho_final = Path(caminho_saida) / subpasta / nome_arquivo
    caminho_final.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(caminho_final))
    logger.info("Consolidado extratos salvo em %s", caminho_final)
    return str(caminho_final)


def montar_caminho_consolidado_extratos(config: dict, mes_referencia: str) -> str:
    """Gera caminho do consolidado de extratos."""
    output_base = config.get("caminhos", {}).get("output_base", "./output")
    output_base = str(Path(output_base).expanduser())
    titular = config.get("titular", "Hugo")
    meses_nomes = config.get("meses", {})

    if mes_referencia and len(mes_referencia) >= 7:
        ano = mes_referencia[:4]
        mes_num = int(mes_referencia[5:7])
    else:
        hoje = datetime.now()
        ano = str(hoje.year)
        mes_num = hoje.month

    nome_mes = meses_nomes.get(mes_num, f"Mes{mes_num:02d}")
    pasta = Path(output_base) / ano / f"{nome_mes}.{ano}"
    nome_arquivo = f"{titular} - Extratos - Consolidado - {nome_mes}.{ano}.xlsx"
    return str(pasta / nome_arquivo)


# ---------------------------------------------------------------------------
# Helpers extratos
# ---------------------------------------------------------------------------

def _todos_lancamentos_extrato(ep: dict) -> list[dict]:
    """Retorna classificados + pendentes + nao_efetivados de um extrato processado."""
    classif = ep.get("classificacao", {})
    return (
        classif.get("classificados", [])
        + classif.get("pendentes", [])
        + classif.get("nao_efetivados", [])
    )


def _lancamentos_efetivados_extrato(ep: dict) -> list[dict]:
    """Retorna apenas lancamentos efetivados (classificados + pendentes)."""
    classif = ep.get("classificacao", {})
    return classif.get("classificados", []) + classif.get("pendentes", [])


# ---------------------------------------------------------------------------
# Abas consolidado extratos
# ---------------------------------------------------------------------------

def _aba_painel_contas(wb: Workbook, extratos: list[dict]) -> None:
    """Aba 'Painel Contas' — uma linha por banco com saldos, totais e alertas."""
    ws = wb.create_sheet(title="Painel Contas")

    headers = [
        "Banco", "Conta", "Saldo Inicial (R$)", "Saldo Final (R$)",
        "Entradas (R$)", "Saídas (R$)", "Não Efetivados (R$)", "Alertas",
    ]
    _escrever_header(ws, headers)

    totais = {
        "saldo_inicial": 0.0, "saldo_final": 0.0,
        "entradas": 0.0, "saidas": 0.0, "nao_efetivados": 0.0,
    }

    row = 2
    for ep in extratos:
        resumo = ep["extrato"].get("resumo", {})
        banco = resumo.get("banco", "")
        conta = resumo.get("conta", "")
        si = resumo.get("saldo_inicial", 0.0)
        sf = resumo.get("saldo_final", 0.0)
        ent = resumo.get("total_entradas", 0.0)
        sai = resumo.get("total_saidas", 0.0)
        ne = resumo.get("total_estornados", 0.0)
        alertas = resumo.get("alertas", [])

        ws.cell(row=row, column=1, value=banco)
        ws.cell(row=row, column=2, value=conta)
        ws.cell(row=row, column=3, value=si).number_format = FMT_BRL
        ws.cell(row=row, column=4, value=sf).number_format = FMT_BRL
        ws.cell(row=row, column=5, value=ent).number_format = FMT_BRL
        ws.cell(row=row, column=6, value=sai).number_format = FMT_BRL
        ws.cell(row=row, column=7, value=ne).number_format = FMT_BRL
        ws.cell(row=row, column=8, value="; ".join(alertas) if alertas else "")

        # Saldo final negativo em vermelho
        if sf < 0:
            ws.cell(row=row, column=4).fill = FILL_VERMELHO

        if row % 2 == 0:
            for col in range(1, 8):
                if ws.cell(row=row, column=col).fill != FILL_VERMELHO:
                    ws.cell(row=row, column=col).fill = FILL_CINZA

        totais["saldo_inicial"] += si
        totais["saldo_final"] += sf
        totais["entradas"] += ent
        totais["saidas"] += sai
        totais["nao_efetivados"] += ne
        row += 1

    # Linha de totais
    ws.cell(row=row, column=1, value="TOTAL CONSOLIDADO").font = FONT_BOLD
    ws.cell(row=row, column=3, value=totais["saldo_inicial"]).number_format = FMT_BRL
    ws.cell(row=row, column=3).font = FONT_BOLD
    ws.cell(row=row, column=4, value=totais["saldo_final"]).number_format = FMT_BRL
    ws.cell(row=row, column=4).font = FONT_BOLD
    ws.cell(row=row, column=5, value=totais["entradas"]).number_format = FMT_BRL
    ws.cell(row=row, column=5).font = FONT_BOLD
    ws.cell(row=row, column=6, value=totais["saidas"]).number_format = FMT_BRL
    ws.cell(row=row, column=6).font = FONT_BOLD
    ws.cell(row=row, column=7, value=totais["nao_efetivados"]).number_format = FMT_BRL
    ws.cell(row=row, column=7).font = FONT_BOLD

    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = FILL_PREVIA

    _aplicar_estilos(ws, num_colunas=8, num_linhas=row)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["H"].width = 60
    ws.sheet_properties.tabColor = CORES["header"]


def _aba_fluxo_consolidado(wb: Workbook, extratos: list[dict]) -> None:
    """Aba 'Fluxo Consolidado' — fluxo de caixa unificado por dia."""
    ws = wb.create_sheet(title="Fluxo Consolidado")

    bancos = []
    for ep in extratos:
        banco = ep["extrato"].get("banco", "?")
        bancos.append(banco)

    headers = ["Data"] + bancos + ["Total Dia (R$)", "Saldo Acumulado (R$)"]
    _escrever_header(ws, headers)

    fluxo: dict[str, dict[str, float]] = defaultdict(lambda: {b: 0.0 for b in bancos})

    for ep in extratos:
        banco = ep["extrato"].get("banco", "?")
        for lanc in _lancamentos_efetivados_extrato(ep):
            data = lanc.get("data", "")
            if data:
                fluxo[data][banco] += lanc.get("valor", 0.0)

    saldo = sum(ep["extrato"].get("resumo", {}).get("saldo_inicial", 0.0) for ep in extratos)

    col_total = len(bancos) + 2
    col_saldo = len(bancos) + 3
    row = 2

    # Linha saldo inicial
    ws.cell(row=row, column=1, value="Saldo Inicial").font = FONT_BOLD
    ws.cell(row=row, column=col_saldo, value=saldo).number_format = FMT_BRL
    ws.cell(row=row, column=col_saldo).font = FONT_BOLD
    for col in range(1, col_saldo + 1):
        ws.cell(row=row, column=col).fill = FILL_PREVIA
    row += 1

    for data in sorted(fluxo.keys()):
        dia = fluxo[data]
        total_dia = sum(dia.values())
        saldo += total_dia

        ws.cell(row=row, column=1, value=data)
        for j, banco in enumerate(bancos):
            val = dia[banco]
            cell = ws.cell(row=row, column=2 + j, value=val if val != 0.0 else "")
            if val != 0.0:
                cell.number_format = FMT_BRL
                if val > 0:
                    cell.fill = FILL_VERDE
                else:
                    cell.fill = FILL_VERMELHO

        ws.cell(row=row, column=col_total, value=total_dia).number_format = FMT_BRL
        cell_saldo = ws.cell(row=row, column=col_saldo, value=round(saldo, 2))
        cell_saldo.number_format = FMT_BRL

        if saldo < 0:
            cell_saldo.fill = FILL_VERMELHO
            cell_saldo.font = Font(bold=True, size=11, color="CC0000")

        row += 1

    _aplicar_estilos(ws, num_colunas=col_saldo, num_linhas=row - 1)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions[get_column_letter(col_saldo)].width = 22
    ws.sheet_properties.tabColor = CORES["verde_l"]


def _aba_receitas_despesas(wb: Workbook, extratos: list[dict]) -> None:
    """Aba 'Receitas vs Despesas' — secoes Receitas, Despesas PF, Despesas PJ."""
    ws = wb.create_sheet(title="Receitas vs Despesas")

    receitas: dict[str, list[float]] = defaultdict(list)
    despesas_pf: dict[str, list[float]] = defaultdict(list)
    despesas_pj: dict[str, list[float]] = defaultdict(list)

    for ep in extratos:
        for lanc in _lancamentos_efetivados_extrato(ep):
            if lanc.get("origem") == "pendente":
                continue

            classif = lanc.get("classificacao", "")
            categoria = lanc.get("categoria", "") or "Sem Categoria"
            valor = lanc.get("valor", 0.0)

            if classif == "PF":
                despesas_pf[categoria].append(valor)
            elif classif == "PJ":
                despesas_pj[categoria].append(valor)
            elif valor > 0:
                receitas[categoria].append(valor)

    row = 1
    row = _escrever_secao_categoria(ws, row, "RECEITAS", receitas, FILL_VERDE, FONT_BOLD)
    row += 1
    row = _escrever_secao_categoria(ws, row, "DESPESAS PF", despesas_pf, FILL_AZUL, FONT_BOLD)
    row += 1
    row = _escrever_secao_categoria(ws, row, "DESPESAS PJ", despesas_pj, FILL_VERDE, FONT_BOLD)

    row += 1
    total_rec = sum(sum(v) for v in receitas.values())
    total_pf = sum(sum(v) for v in despesas_pf.values())
    total_pj = sum(sum(v) for v in despesas_pj.values())
    total_geral = total_rec + total_pf + total_pj

    ws.cell(row=row, column=1, value="TOTAL GERAL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=total_geral).number_format = FMT_BRL
    ws.cell(row=row, column=2).font = FONT_BOLD
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = FILL_PREVIA

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.sheet_properties.tabColor = CORES["azul_pj"]


def _aba_relatorio_executivo(
    wb: Workbook, extratos: list[dict], mes_referencia: str,
) -> None:
    """Aba 'Relatorio Executivo' — posicao financeira, maiores movimentacoes, alertas."""
    ws = wb.create_sheet(title="Relatorio Executivo")

    ws.merge_cells("A1:E1")
    cell = ws.cell(
        row=1, column=1,
        value="Relatório Executivo — Posição Financeira Consolidada",
    )
    cell.font = Font(bold=True, size=14, color=CORES["header"])
    cell.alignment = ALIGN_CENTER

    ws.merge_cells("A2:E2")
    ws.cell(row=2, column=1, value=f"Referência: {mes_referencia}").alignment = ALIGN_CENTER

    # --- Posicao financeira ---
    row = 4
    ws.cell(row=row, column=1, value="Posição Financeira por Conta").font = Font(bold=True, size=12)
    row += 1
    _escrever_header_em_row(ws, row, ["Banco", "Saldo Inicial", "Entradas", "Saídas", "Saldo Final"])
    row += 1

    total_si = total_sf = total_ent = total_sai = 0.0

    for ep in extratos:
        resumo = ep["extrato"].get("resumo", {})
        banco = resumo.get("banco", "")
        si = resumo.get("saldo_inicial", 0.0)
        sf = resumo.get("saldo_final", 0.0)
        ent = resumo.get("total_entradas", 0.0)
        sai = resumo.get("total_saidas", 0.0)

        ws.cell(row=row, column=1, value=banco).font = FONT_BOLD
        ws.cell(row=row, column=2, value=si).number_format = FMT_BRL
        ws.cell(row=row, column=3, value=ent).number_format = FMT_BRL
        ws.cell(row=row, column=4, value=sai).number_format = FMT_BRL
        ws.cell(row=row, column=5, value=sf).number_format = FMT_BRL
        if sf < 0:
            ws.cell(row=row, column=5).fill = FILL_VERMELHO

        total_si += si
        total_sf += sf
        total_ent += ent
        total_sai += sai
        row += 1

    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = FILL_PREVIA
    ws.cell(row=row, column=1, value="CONSOLIDADO").font = FONT_BOLD
    ws.cell(row=row, column=2, value=total_si).number_format = FMT_BRL
    ws.cell(row=row, column=2).font = FONT_BOLD
    ws.cell(row=row, column=3, value=total_ent).number_format = FMT_BRL
    ws.cell(row=row, column=3).font = FONT_BOLD
    ws.cell(row=row, column=4, value=total_sai).number_format = FMT_BRL
    ws.cell(row=row, column=4).font = FONT_BOLD
    ws.cell(row=row, column=5, value=total_sf).number_format = FMT_BRL
    ws.cell(row=row, column=5).font = FONT_BOLD
    row += 2

    # --- Maiores entradas ---
    ws.cell(row=row, column=1, value="Maiores Entradas do Período").font = Font(bold=True, size=12)
    row += 1
    _escrever_header_em_row(ws, row, ["Banco", "Data", "Descrição", "Valor (R$)", ""])
    row += 1

    todas_entradas = []
    for ep in extratos:
        banco = ep["extrato"].get("banco", "?")
        for lanc in _lancamentos_efetivados_extrato(ep):
            if lanc.get("valor", 0.0) > 0:
                todas_entradas.append({**lanc, "_banco": banco})
    todas_entradas.sort(key=lambda l: l.get("valor", 0.0), reverse=True)

    for lanc in todas_entradas[:10]:
        ws.cell(row=row, column=1, value=lanc["_banco"])
        ws.cell(row=row, column=2, value=lanc.get("data", ""))
        ws.cell(row=row, column=3, value=lanc.get("descricao", ""))
        ws.cell(row=row, column=4, value=lanc.get("valor", 0.0)).number_format = FMT_BRL
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = FILL_VERDE
        row += 1
    row += 1

    # --- Maiores saidas ---
    ws.cell(row=row, column=1, value="Maiores Saídas do Período").font = Font(bold=True, size=12)
    row += 1
    _escrever_header_em_row(ws, row, ["Banco", "Data", "Descrição", "Valor (R$)", ""])
    row += 1

    todas_saidas = []
    for ep in extratos:
        banco = ep["extrato"].get("banco", "?")
        for lanc in _lancamentos_efetivados_extrato(ep):
            if lanc.get("valor", 0.0) < 0:
                todas_saidas.append({**lanc, "_banco": banco})
    todas_saidas.sort(key=lambda l: l.get("valor", 0.0))

    for lanc in todas_saidas[:10]:
        ws.cell(row=row, column=1, value=lanc["_banco"])
        ws.cell(row=row, column=2, value=lanc.get("data", ""))
        ws.cell(row=row, column=3, value=lanc.get("descricao", ""))
        ws.cell(row=row, column=4, value=lanc.get("valor", 0.0)).number_format = FMT_BRL
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = FILL_VERMELHO
        row += 1
    row += 1

    # --- Alertas consolidados ---
    ws.cell(row=row, column=1, value="Alertas").font = Font(bold=True, size=12, color="CC0000")
    row += 1

    tem_alertas = False
    for ep in extratos:
        resumo = ep["extrato"].get("resumo", {})
        banco = resumo.get("banco", "")
        alertas = resumo.get("alertas", [])
        for alerta in alertas:
            ws.cell(row=row, column=1, value=f"{banco}: {alerta}")
            ws.cell(row=row, column=1).font = Font(size=11, color="CC0000")
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = FILL_ALERTA
            row += 1
            tem_alertas = True

    if not tem_alertas:
        ws.cell(row=row, column=1, value="Nenhum alerta no período.")
        row += 1

    # --- Saldo disponivel total ---
    row += 1
    ws.cell(row=row, column=1, value="Saldo Disponível Total").font = Font(bold=True, size=14)
    ws.cell(row=row, column=2, value=total_sf).number_format = FMT_BRL
    ws.cell(row=row, column=2).font = Font(bold=True, size=14)
    fill_saldo = FILL_VERDE if total_sf >= 0 else FILL_VERMELHO
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = fill_saldo

    _aplicar_estilos(ws, num_colunas=5, num_linhas=row)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.sheet_properties.tabColor = CORES["verde_pj"]
