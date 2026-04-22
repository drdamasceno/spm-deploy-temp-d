# spm-faturas: Extrator de extratos bancários (XLSX e PDF)

from __future__ import annotations

import re
import logging
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

import openpyxl
import pdfplumber

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# API publica
# ---------------------------------------------------------------------------

def extrair_extrato_bb(caminho_xlsx: str) -> dict:
    """Entry point para extratos BB em XLSX.

    Retorna dict com 'banco', 'conta', 'mes_referencia',
    'resumo' e 'lancamentos'.
    """
    caminho = Path(caminho_xlsx)
    if not caminho.exists():
        raise FileNotFoundError(f"XLSX nao encontrado: {caminho_xlsx}")

    linhas = _ler_linhas_xlsx(str(caminho))
    lancamentos, info_controle = _filtrar_linhas_controle(linhas)
    lancamentos = _detectar_estornos(lancamentos)

    mes_referencia = _inferir_mes_referencia(lancamentos)
    alertas = _gerar_alertas(lancamentos, info_controle)

    resumo = _calcular_resumo(
        lancamentos,
        info_controle["saldo_inicial"],
        info_controle["saldo_final"],
        mes_referencia,
        alertas,
    )

    return {
        "banco": "BB",
        "conta": "corrente",
        "mes_referencia": mes_referencia,
        "resumo": resumo,
        "lancamentos": lancamentos,
    }


# ---------------------------------------------------------------------------
# Leitura do XLSX
# ---------------------------------------------------------------------------

def _ler_linhas_xlsx(caminho: str) -> list[dict]:
    """Le XLSX com openpyxl, retorna lista de dicts crus (1 por linha).

    Pula header (row 1). Cada dict tem:
    data_raw, lancamento, detalhes, documento, valor_raw, tipo_lancamento.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active

    linhas = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or len(row) < 6:
            continue

        data_raw = str(row[0] or "").strip()
        lancamento = str(row[1] or "").strip()
        detalhes = str(row[2] or "").strip()
        documento = str(row[3] or "").strip()
        valor_raw = str(row[4] or "").strip()
        tipo_lancamento = str(row[5] or "").strip()

        linhas.append({
            "data_raw": data_raw,
            "lancamento": lancamento,
            "detalhes": detalhes,
            "documento": documento,
            "valor_raw": valor_raw,
            "tipo_lancamento": tipo_lancamento,
            "row": i,
        })

    wb.close()
    return linhas


# ---------------------------------------------------------------------------
# Filtragem de linhas de controle
# ---------------------------------------------------------------------------

def _filtrar_linhas_controle(linhas: list[dict]) -> tuple[list[dict], dict]:
    """Separa lancamentos reais das linhas de controle.

    Remove: data 00/00/0000 (Saldo do dia), S A L D O, Saldo Anterior.
    Retorna (lancamentos_filtrados, info_controle) onde info_controle
    contem saldo_inicial, saldo_final, e saldos_diarios.
    """
    info = {
        "saldo_inicial": 0.0,
        "saldo_final": 0.0,
        "saldos_diarios": [],
    }

    lancamentos = []

    for linha in linhas:
        lancamento = linha["lancamento"]

        # Saldo Anterior — primeira linha real do extrato
        if lancamento == "Saldo Anterior":
            info["saldo_inicial"] = _parse_valor_br(linha["valor_raw"])
            continue

        # S A L D O — ultima linha do extrato
        if lancamento == "S A L D O":
            info["saldo_final"] = _parse_valor_br(linha["valor_raw"])
            continue

        # Saldo do dia — linhas com data 00/00/0000
        if linha["data_raw"] == "00/00/0000":
            saldo_dia = _parse_valor_br(linha["valor_raw"])
            info["saldos_diarios"].append(saldo_dia)
            continue

        # Lancamento real — montar dict de saida
        valor = _parse_valor_br(linha["valor_raw"])
        tipo = _classificar_tipo(lancamento, linha["tipo_lancamento"])

        lancamentos.append({
            "data": _parse_data_br(linha["data_raw"]),
            "descricao": lancamento,
            "detalhes": detalhes if (detalhes := linha["detalhes"]) else "",
            "documento": linha["documento"],
            "valor": valor,
            "tipo": tipo,
            "efetivado": True,
            "categoria": "",
            "classificacao": "",
        })

    return lancamentos, info


# ---------------------------------------------------------------------------
# Deteccao de estornos
# ---------------------------------------------------------------------------

def _detectar_estornos(lancamentos: list[dict]) -> list[dict]:
    """Marca pares de estorno: mesmo N documento + descricao de estorno + valores opostos.

    Regra unica: agrupa por documento, emparelha cada debito (valor < 0)
    com um estorno (valor > 0, descricao contem 'estorno') do MESMO documento
    cujo valor seja exatamente oposto (tolerancia R$ 0.01).

    Nao cruza documentos diferentes. Nao usa proximidade temporal.
    """
    por_documento: dict[str, list[int]] = defaultdict(list)
    for i, lanc in enumerate(lancamentos):
        doc = lanc["documento"]
        if doc and doc.strip():
            por_documento[doc].append(i)

    for doc, indices in por_documento.items():
        if len(indices) < 2:
            continue

        debitos = []
        estornos = []
        for idx in indices:
            lanc = lancamentos[idx]
            if lanc["valor"] > 0 and "estorno" in lanc["descricao"].lower():
                estornos.append(idx)
            elif lanc["valor"] < 0:
                debitos.append(idx)

        if not debitos or not estornos:
            continue

        estornos_usados: set[int] = set()

        for idx_deb in debitos:
            deb_valor = lancamentos[idx_deb]["valor"]

            for idx_est in estornos:
                if idx_est in estornos_usados:
                    continue
                if abs(deb_valor + lancamentos[idx_est]["valor"]) > 0.01:
                    continue

                lancamentos[idx_deb]["efetivado"] = False
                lancamentos[idx_est]["efetivado"] = False
                estornos_usados.add(idx_est)
                break

    return lancamentos


# ---------------------------------------------------------------------------
# Classificacao de tipo
# ---------------------------------------------------------------------------

KEYWORDS_ENCARGO = [
    "I.O.F.", "IOF", "JUROS", "COBRANÇA DE JUROS",
    "ENCARGO", "MULTA", "MORA",
]

KEYWORDS_TARIFA = [
    "TARIFA", "PACOTE DE SERVIÇOS", "ANUIDADE",
]


def _classificar_tipo(lancamento: str, tipo_lancamento: str) -> str:
    """Retorna tipo semantico: entrada | saida | encargo | tarifa.

    Usa campo 'Lancamento' para detectar encargos (IOF, Juros)
    e tarifas (Tarifa Pacote). Default: entrada/saida pelo Tipo Lancamento.
    """
    lanc_upper = lancamento.upper()

    for kw in KEYWORDS_TARIFA:
        if kw in lanc_upper:
            return "tarifa"

    for kw in KEYWORDS_ENCARGO:
        if kw in lanc_upper:
            return "encargo"

    if tipo_lancamento == "Entrada":
        return "entrada"

    return "saida"


# ---------------------------------------------------------------------------
# Alertas
# ---------------------------------------------------------------------------

def _gerar_alertas(lancamentos: list[dict], info_controle: dict) -> list[str]:
    """Gera alertas de valor para o usuario.

    Detecta:
    - Grupos de tentativas nao efetivadas (debito + estorno repetidos)
    - Saldo negativo (primeiro dia em que ficou negativo)
    """
    alertas = []

    # --- Agrupar nao-efetivados por descricao ---
    nao_efetivados: dict[str, int] = defaultdict(int)
    for lanc in lancamentos:
        if not lanc["efetivado"] and lanc["valor"] < 0:
            # Normalizar descricao para agrupamento
            desc = lanc["descricao"]
            nao_efetivados[desc] += 1

    NOMES_AMIGAVEIS = {
        "FIES JRS/AMORTIZACAO": "FIES",
        "Pagto cartão crédito": "Cartão BB",
    }

    for desc, qtd in sorted(nao_efetivados.items(), key=lambda x: -x[1]):
        nome = NOMES_AMIGAVEIS.get(desc, desc)
        if qtd == 1:
            alertas.append(f"{nome}: 1 tentativa de débito não efetivada")
        else:
            alertas.append(f"{nome}: {qtd} tentativas de débito não efetivadas")

    # --- Saldo negativo ---
    saldo = info_controle["saldo_inicial"]
    primeiro_negativo = None

    for lanc in lancamentos:
        if lanc["efetivado"]:
            saldo += lanc["valor"]
            if saldo < 0 and primeiro_negativo is None:
                primeiro_negativo = lanc["data"]

    if primeiro_negativo:
        data_fmt = _formatar_data_br(primeiro_negativo)
        alertas.append(f"Saldo ficou negativo em {data_fmt}")

    return alertas


# ---------------------------------------------------------------------------
# Resumo
# ---------------------------------------------------------------------------

def _calcular_resumo(
    lancamentos: list[dict],
    saldo_inicial: float,
    saldo_final: float,
    mes_referencia: str,
    alertas: list[str],
) -> dict:
    """Calcula resumo: totais de entrada, saida, estornados."""
    total_entradas = 0.0
    total_saidas = 0.0
    total_estornados = 0.0

    for lanc in lancamentos:
        if not lanc["efetivado"]:
            if lanc["valor"] < 0:
                total_estornados += abs(lanc["valor"])
            continue

        if lanc["valor"] > 0:
            total_entradas += lanc["valor"]
        else:
            total_saidas += lanc["valor"]

    return {
        "banco": "BB",
        "conta": "corrente",
        "mes_referencia": mes_referencia,
        "saldo_inicial": saldo_inicial,
        "saldo_final": saldo_final,
        "total_entradas": round(total_entradas, 2),
        "total_saidas": round(total_saidas, 2),
        "total_estornados": round(total_estornados, 2),
        "alertas": alertas,
    }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_valor_br(texto: str) -> float:
    """Converte '1.234,56' ou '-1.234,56' para float."""
    texto = texto.strip()
    if not texto:
        return 0.0
    negativo = texto.startswith("-") or texto.endswith("-")
    texto = texto.replace("-", "").strip()
    texto = texto.replace(".", "").replace(",", ".")
    valor = float(texto)
    return -valor if negativo else valor


def _parse_data_br(texto: str) -> str:
    """Converte 'DD/MM/YYYY' para 'YYYY-MM-DD'."""
    texto = texto.strip()
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", texto)
    if not m:
        return ""
    dia, mes, ano = m.group(1), m.group(2), m.group(3)
    return f"{ano}-{mes}-{dia}"


def _parse_date_obj(data_iso: str) -> datetime:
    """Converte 'YYYY-MM-DD' para datetime."""
    return datetime.strptime(data_iso, "%Y-%m-%d")


def _formatar_data_br(data_iso: str) -> str:
    """Converte 'YYYY-MM-DD' para 'DD/MM/YYYY'."""
    if len(data_iso) != 10:
        return data_iso
    ano, mes, dia = data_iso.split("-")
    return f"{dia}/{mes}/{ano}"


def _inferir_mes_referencia(lancamentos: list[dict]) -> str:
    """Infere mes de referencia a partir das datas dos lancamentos.

    Usa a moda (mes mais frequente) entre os lancamentos efetivados.
    """
    contagem: dict[str, int] = defaultdict(int)
    for lanc in lancamentos:
        data = lanc.get("data", "")
        if len(data) >= 7:
            contagem[data[:7]] += 1

    if not contagem:
        return ""

    return max(contagem, key=contagem.get)


# ===========================================================================
# ITAU — Extrato PDF
# ===========================================================================

COL_VALOR_ITAU = 480  # x-boundary: valor (< 480) vs saldo (>= 480)


def extrair_extrato_itau(caminho_pdf: str) -> dict:
    """Entry point para extratos Itau em PDF.

    Retorna dict com 'banco', 'conta', 'mes_referencia',
    'resumo' e 'lancamentos'.
    """
    caminho = Path(caminho_pdf)
    if not caminho.exists():
        raise FileNotFoundError(f"PDF nao encontrado: {caminho_pdf}")

    with pdfplumber.open(caminho) as pdf:
        # Extrair info do cabecalho (pagina 1)
        texto_pag1 = pdf.pages[0].extract_text() or ""
        info = _extrair_info_itau(texto_pag1)

        # Extrair lancamentos usando word positions (todas as paginas)
        lancamentos_raw = []
        saldos_diarios = []
        for pag in pdf.pages:
            lancs, saldos = _extrair_linhas_itau(pag)
            lancamentos_raw.extend(lancs)
            saldos_diarios.extend(saldos)

    # Ordenar cronologicamente (PDF Itau vem em ordem reversa)
    lancamentos_raw.sort(key=lambda l: l["data"])
    saldos_diarios.sort(key=lambda s: s["data"])

    # Saldo inicial = primeiro SALDO DO DIA, saldo final = ultimo
    saldo_inicial = saldos_diarios[0]["valor"] if saldos_diarios else 0.0
    saldo_final = saldos_diarios[-1]["valor"] if saldos_diarios else 0.0

    # Classificar tipos
    lancamentos = []
    for raw in lancamentos_raw:
        tipo = _classificar_tipo_itau(raw["descricao"], raw["valor"])
        lancamentos.append({
            "data": raw["data"],
            "descricao": raw["descricao"],
            "detalhes": "",
            "documento": "",
            "valor": raw["valor"],
            "tipo": tipo,
            "efetivado": True,  # Itau nao tem logica de estornos
            "categoria": "",
            "classificacao": "",
        })

    mes_referencia = _inferir_mes_referencia(lancamentos)

    # Alertas
    alertas = _gerar_alertas_itau(lancamentos, saldo_inicial, saldos_diarios)

    resumo = {
        "banco": "Itau",
        "conta": info.get("conta", "corrente"),
        "mes_referencia": mes_referencia,
        "saldo_inicial": saldo_inicial,
        "saldo_final": saldo_final,
        "total_entradas": round(sum(l["valor"] for l in lancamentos if l["valor"] > 0), 2),
        "total_saidas": round(sum(l["valor"] for l in lancamentos if l["valor"] < 0), 2),
        "total_estornados": 0.0,
        "alertas": alertas,
    }

    return {
        "banco": "Itau",
        "conta": info.get("conta", "corrente"),
        "mes_referencia": mes_referencia,
        "resumo": resumo,
        "lancamentos": lancamentos,
    }


def _extrair_info_itau(texto: str) -> dict:
    """Extrai titular, agencia e conta do cabecalho/rodape do extrato Itau."""
    info = {"titular": "", "agencia": "", "conta": "corrente"}

    # "HUGO FERNANDES DAMASCENO 100.022.976-98 agência: 4081 conta: 010791-6"
    m = re.search(r'agência:\s*(\d+)', texto)
    if m:
        info["agencia"] = m.group(1)

    m = re.search(r'conta:\s*([\d-]+)', texto)
    if m:
        info["conta"] = m.group(1)

    m = re.search(r'^([A-Z][A-Z\s]+?)\s+\d{3}\.\d{3}\.\d{3}', texto, re.MULTILINE)
    if m:
        info["titular"] = m.group(1).strip()

    return info


def _extrair_linhas_itau(pagina) -> tuple[list[dict], list[dict]]:
    """Extrai lancamentos e saldos diarios de uma pagina do extrato Itau.

    Usa word positions para distinguir coluna valor (x < 480) de saldo (x >= 480).
    Retorna (lancamentos, saldos_diarios).
    """
    words = pagina.extract_words(
        keep_blank_chars=True, x_tolerance=2, y_tolerance=2
    )

    # Agrupar words por linha (top arredondado)
    lines_by_y: dict[float, list] = defaultdict(list)
    for w in words:
        lines_by_y[round(w['top'], 0)].append(w)

    lancamentos = []
    saldos = []

    RE_DATA = re.compile(r'^\d{2}/\d{2}/\d{4}$')
    RE_VALOR = re.compile(r'^-?[\d.,]+$')

    for y in sorted(lines_by_y.keys()):
        ws = sorted(lines_by_y[y], key=lambda w: w['x0'])

        # Primeira word deve ser uma data DD/MM/YYYY
        if not ws or not RE_DATA.match(ws[0]['text'].strip()):
            continue

        data_raw = ws[0]['text'].strip()
        data = _parse_data_br(data_raw)

        # Coletar descricao (words entre data e valor/saldo)
        desc_parts = []
        valor_encontrado = None
        saldo_encontrado = None

        for w in ws[1:]:
            txt = w['text'].strip()
            if RE_VALOR.match(txt):
                val = _parse_valor_br(txt)
                if w['x0'] >= COL_VALOR_ITAU:
                    saldo_encontrado = val
                else:
                    valor_encontrado = val
            else:
                desc_parts.append(txt)

        descricao = ' '.join(desc_parts).strip()

        # SALDO DO DIA → saldo diario (ignorar como lancamento)
        if 'SALDO DO DIA' in descricao.upper():
            valor_saldo = saldo_encontrado if saldo_encontrado is not None else valor_encontrado
            if valor_saldo is not None:
                saldos.append({"data": data, "valor": valor_saldo})
            continue

        # Lancamento real — deve ter valor na coluna valor
        if valor_encontrado is not None:
            lancamentos.append({
                "data": data,
                "descricao": descricao,
                "valor": valor_encontrado,
            })

    return lancamentos, saldos


KEYWORDS_ENCARGO_ITAU = ["IOF", "JUROS", "ENCARGO", "MULTA", "MORA"]
KEYWORDS_TARIFA_ITAU = ["TARIFA", "ANUIDADE", "PACOTE"]


def _classificar_tipo_itau(descricao: str, valor: float) -> str:
    """Retorna tipo semantico para lancamento de extrato Itau."""
    desc_upper = descricao.upper()

    for kw in KEYWORDS_TARIFA_ITAU:
        if kw in desc_upper:
            return "tarifa"

    for kw in KEYWORDS_ENCARGO_ITAU:
        if kw in desc_upper:
            return "encargo"

    return "entrada" if valor > 0 else "saida"


def _gerar_alertas_itau(
    lancamentos: list[dict],
    saldo_inicial: float,
    saldos_diarios: list[dict],
) -> list[str]:
    """Gera alertas para extrato Itau."""
    alertas = []

    # IOF cobrado
    iof_total = sum(
        abs(l["valor"]) for l in lancamentos
        if "IOF" in l["descricao"].upper()
    )
    if iof_total > 0:
        alertas.append(f"IOF cobrado: R$ {iof_total:,.2f}")

    # Saldo negativo
    for s in saldos_diarios:
        if s["valor"] < 0:
            data_fmt = _formatar_data_br(s["data"])
            alertas.append(f"Saldo ficou negativo em {data_fmt}")
            break

    return alertas


# ===========================================================================
# UNICRED — Extrato PDF
# ===========================================================================

def extrair_extrato_unicred(caminho_pdf: str) -> dict:
    """Entry point para extratos Unicred em PDF.

    Retorna dict com 'banco', 'conta', 'mes_referencia',
    'resumo' e 'lancamentos'.
    """
    caminho = Path(caminho_pdf)
    if not caminho.exists():
        raise FileNotFoundError(f"PDF nao encontrado: {caminho_pdf}")

    with pdfplumber.open(caminho) as pdf:
        texto_completo = ""
        for pag in pdf.pages:
            texto_completo += (pag.extract_text() or "") + "\n"

    info = _extrair_info_unicred_extrato(texto_completo)
    saldo_inicial = _extrair_saldo_anterior_unicred_extrato(texto_completo)
    saldo_final = _extrair_saldo_final_unicred_extrato(texto_completo)

    lancamentos_raw = _extrair_lancamentos_unicred_extrato(texto_completo)

    # Montar lancamentos com tipo e efetivado
    lancamentos = []
    for raw in lancamentos_raw:
        tipo = _classificar_tipo_unicred_extrato(raw["descricao"], raw["valor"])
        lancamentos.append({
            "data": raw["data"],
            "descricao": raw["descricao"],
            "detalhes": raw.get("detalhes", ""),
            "documento": raw["documento"],
            "valor": raw["valor"],
            "tipo": tipo,
            "efetivado": True,
            "categoria": "",
            "classificacao": "",
        })

    # Detectar estornos (reutiliza logica do BB)
    lancamentos = _detectar_estornos(lancamentos)

    mes_referencia = _inferir_mes_referencia(lancamentos)
    alertas = _gerar_alertas_unicred_extrato(lancamentos, saldo_inicial)

    resumo = {
        "banco": "Unicred",
        "conta": info.get("conta", "corrente"),
        "mes_referencia": mes_referencia,
        "saldo_inicial": saldo_inicial,
        "saldo_final": saldo_final,
        "total_entradas": round(sum(
            l["valor"] for l in lancamentos if l["valor"] > 0 and l["efetivado"]
        ), 2),
        "total_saidas": round(sum(
            l["valor"] for l in lancamentos if l["valor"] < 0 and l["efetivado"]
        ), 2),
        "total_estornados": round(sum(
            abs(l["valor"]) for l in lancamentos if not l["efetivado"] and l["valor"] < 0
        ), 2),
        "alertas": alertas,
    }

    return {
        "banco": "Unicred",
        "conta": info.get("conta", "corrente"),
        "mes_referencia": mes_referencia,
        "resumo": resumo,
        "lancamentos": lancamentos,
    }


def _extrair_info_unicred_extrato(texto: str) -> dict:
    """Extrai conta e cliente do cabecalho do extrato Unicred."""
    info = {"conta": "corrente", "titular": ""}

    m = re.search(r'CONTA:\s*(\d+)', texto)
    if m:
        info["conta"] = m.group(1)

    m = re.search(r'CLIENTE:\s*(.+)', texto)
    if m:
        info["titular"] = m.group(1).strip()

    return info


def _extrair_saldo_anterior_unicred_extrato(texto: str) -> float:
    """Extrai saldo anterior da linha 'SALDO ANTERIOR: valor'."""
    m = re.search(r'SALDO ANTERIOR:\s+([\d.,]+)', texto)
    if m:
        return _parse_valor_br(m.group(1))
    return 0.0


def _extrair_saldo_final_unicred_extrato(texto: str) -> float:
    """Extrai saldo final da linha 'Saldo em DD/MM/YYYY valor'."""
    m = re.search(r'Saldo em \d{2}/\d{2}/\d{4}\s+(-?[\d.,]+)', texto)
    if m:
        return _parse_valor_br(m.group(1))
    return 0.0


def _extrair_lancamentos_unicred_extrato(texto: str) -> list[dict]:
    """Extrai lancamentos do texto do extrato Unicred.

    Formato por linha (separador (cid:9)):
    DD/MM/YYYY  ID_DOC  (cid:9) HISTORICO (cid:9) VALOR  SALDO

    Caso especial PIX: valor na mesma linha apos historico, saldo na linha seguinte.
    """
    linhas = texto.split('\n')
    lancamentos = []

    RE_LINHA = re.compile(
        r'^(\d{2}/\d{2}/\d{4})\s+(.+?)\(cid:9\)(.+?)\(cid:9\)\s*(.*)$'
    )
    RE_VALOR_NUM = re.compile(r'-?[\d.,]+')

    STOP_MARKERS = ['Saldo em ', 'Limite de Cr', 'Saldo Bloqueado', 'CPMF', 'Juros Cheque']

    for linha in linhas:
        linha = linha.strip()
        if not linha:
            continue

        if any(linha.startswith(marker) for marker in STOP_MARKERS):
            break

        m = RE_LINHA.match(linha)
        if not m:
            continue

        data_raw = m.group(1)
        doc_raw = m.group(2).strip()
        historico_raw = m.group(3).strip()
        valores_raw = m.group(4).strip()

        # Limpar (cid:9) residuais
        doc_raw = re.sub(r'\(cid:9\)', '', doc_raw).strip()
        historico_raw = re.sub(r'\(cid:9\)', '', historico_raw).strip()

        # Extrair valor do lancamento (primeiro numero na coluna valores)
        valores_match = RE_VALOR_NUM.findall(valores_raw)
        if valores_match:
            valor = _parse_valor_br(valores_match[0])
        else:
            # PIX: valor pode estar no final do historico (apos ultimo texto)
            # Ex: "CRED RECEBIMENTO PIX(...)  54.837,61 "
            # Nesse caso valores_raw esta vazio, valor esta colado no historico
            all_nums = RE_VALOR_NUM.findall(historico_raw)
            if all_nums:
                valor = _parse_valor_br(all_nums[-1])
            else:
                continue

        # Detalhes: conteudo entre parenteses no historico
        detalhes = ""
        m_det = re.search(r'\((.+?)\)', historico_raw)
        if m_det:
            detalhes = m_det.group(1)

        # Limpar descricao: remover detalhes entre parenteses
        descricao = re.sub(r'\(.+?\)', '', historico_raw).strip()

        lancamentos.append({
            "data": _parse_data_br(data_raw),
            "descricao": descricao,
            "detalhes": detalhes,
            "documento": doc_raw,
            "valor": valor,
        })

    return lancamentos


KEYWORDS_ENCARGO_UNICRED_EXT = ["IOF", "JUROS", "ENCARGO", "MULTA"]
KEYWORDS_TARIFA_UNICRED_EXT = ["PF PCT PADRONIZADO", "TARIFA", "ANUIDADE"]


def _classificar_tipo_unicred_extrato(descricao: str, valor: float) -> str:
    """Retorna tipo semantico para lancamento de extrato Unicred."""
    desc_upper = descricao.upper()

    for kw in KEYWORDS_TARIFA_UNICRED_EXT:
        if kw in desc_upper:
            return "tarifa"

    for kw in KEYWORDS_ENCARGO_UNICRED_EXT:
        if kw in desc_upper:
            return "encargo"

    return "entrada" if valor > 0 else "saida"


def _gerar_alertas_unicred_extrato(
    lancamentos: list[dict], saldo_inicial: float,
) -> list[str]:
    """Gera alertas para extrato Unicred."""
    alertas = []

    # Tentativas nao efetivadas
    nao_efetivados: dict[str, list[str]] = defaultdict(list)
    for lanc in lancamentos:
        if not lanc["efetivado"] and lanc["valor"] < 0:
            desc = lanc["descricao"]
            nao_efetivados[desc].append(lanc["data"])

    for desc, datas in nao_efetivados.items():
        for data in datas:
            data_fmt = _formatar_data_br(data)
            alertas.append(f"{desc}: tentativa de débito não efetivada em {data_fmt}")

    # Saldo negativo
    saldo = saldo_inicial
    for lanc in lancamentos:
        if lanc["efetivado"]:
            saldo += lanc["valor"]
            if saldo < 0:
                data_fmt = _formatar_data_br(lanc["data"])
                alertas.append(f"Saldo ficou negativo em {data_fmt}")
                break

    return alertas


# ===========================================================================
# SISPRIME — Extrato PDF
# ===========================================================================

# Keywords que indicam credito (entrada) no historico Sisprime
CREDITO_KEYWORDS_SISPRIME = ["Crédito Pix", "Estorno Débito"]


def extrair_extrato_sisprime(caminho_pdf: str) -> dict:
    """Entry point para extratos Sisprime em PDF.

    Retorna dict com 'banco', 'conta', 'mes_referencia',
    'resumo' e 'lancamentos'.
    """
    caminho = Path(caminho_pdf)
    if not caminho.exists():
        raise FileNotFoundError(f"PDF nao encontrado: {caminho_pdf}")

    with pdfplumber.open(caminho) as pdf:
        texto_completo = ""
        for pag in pdf.pages:
            texto_completo += (pag.extract_text() or "") + "\n"

    info = _extrair_info_sisprime(texto_completo)
    saldo_inicial = _extrair_saldo_anterior_sisprime(texto_completo)
    saldo_final = _extrair_saldo_final_sisprime(texto_completo)

    lancamentos_raw = _extrair_lancamentos_sisprime_extrato(texto_completo)

    # Montar lancamentos com tipo e efetivado
    lancamentos = []
    for raw in lancamentos_raw:
        tipo = _classificar_tipo_sisprime_extrato(raw["historico"], raw["valor"])
        lancamentos.append({
            "data": raw["data"],
            "descricao": raw["descricao"],
            "detalhes": raw.get("detalhes", ""),
            "documento": raw["documento"],
            "valor": raw["valor"],
            "tipo": tipo,
            "efetivado": True,
            "categoria": "",
            "classificacao": "",
        })

    # Detectar estornos (reutiliza logica do BB)
    lancamentos = _detectar_estornos(lancamentos)

    mes_referencia = _inferir_mes_referencia(lancamentos)
    alertas = _gerar_alertas_sisprime(lancamentos, saldo_inicial)

    resumo = {
        "banco": "Sisprime",
        "conta": info.get("conta", "corrente"),
        "mes_referencia": mes_referencia,
        "saldo_inicial": saldo_inicial,
        "saldo_final": saldo_final,
        "total_entradas": round(sum(
            l["valor"] for l in lancamentos if l["valor"] > 0 and l["efetivado"]
        ), 2),
        "total_saidas": round(sum(
            l["valor"] for l in lancamentos if l["valor"] < 0 and l["efetivado"]
        ), 2),
        "total_estornados": round(sum(
            abs(l["valor"]) for l in lancamentos if not l["efetivado"] and l["valor"] < 0
        ), 2),
        "alertas": alertas,
    }

    return {
        "banco": "Sisprime",
        "conta": info.get("conta", "corrente"),
        "mes_referencia": mes_referencia,
        "resumo": resumo,
        "lancamentos": lancamentos,
    }


def _extrair_info_sisprime(texto: str) -> dict:
    """Extrai conta e titular do cabecalho do extrato Sisprime."""
    info = {"conta": "corrente", "titular": ""}

    m = re.search(r'Conta:\s*(\d+)', texto)
    if m:
        info["conta"] = m.group(1)

    m = re.search(r'Nome:\s*(.+)', texto)
    if m:
        info["titular"] = m.group(1).strip()

    return info


def _extrair_saldo_anterior_sisprime(texto: str) -> float:
    """Extrai saldo anterior: 'Saldo Anterior: R$ 12.546,09'."""
    m = re.search(r'Saldo Anterior:\s*R\$\s*([\d.,]+)', texto)
    if m:
        return _parse_valor_br(m.group(1))
    return 0.0


def _extrair_saldo_final_sisprime(texto: str) -> float:
    """Extrai saldo final: 'Saldo: R$ 18.911,27' (apos 'Posição em')."""
    m = re.search(r'Posição em.*?\nSaldo:\s*R\$\s*([\d.,]+)', texto, re.DOTALL)
    if m:
        return _parse_valor_br(m.group(1))
    return 0.0


def _extrair_lancamentos_sisprime_extrato(texto: str) -> list[dict]:
    """Extrai lancamentos do texto do extrato Sisprime.

    Formato por linha:
    DD/MM/YYYY DOCUMENTO HISTORICO [DESCRICAO] R$ VALOR [-]R$ SALDO

    Cada linha tem exatamente 2 valores R$: transacao e saldo corrente.
    Debito/credito determinado pelo historico (Crédito Pix = entrada, etc).
    """
    linhas = texto.split('\n')
    lancamentos = []

    # Regex: linha comecando com data
    RE_DATA_START = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+')
    # Regex: extrair todos os R$ valores da linha
    RE_VALOR_RS = re.compile(r'-?R\$\s*([\d.,]+)')

    STOP_MARKERS = ['Posição em ']
    SKIP_LINES = ['Data Documento', 'Lançamentos Saldo', 'OS DADOS ACIMA']

    for linha in linhas:
        linha = linha.strip()
        if not linha:
            continue

        if any(linha.startswith(marker) for marker in STOP_MARKERS):
            break

        if any(linha.startswith(skip) for skip in SKIP_LINES):
            continue

        m_data = RE_DATA_START.match(linha)
        if not m_data:
            continue

        data_raw = m_data.group(1)

        # Extrair valores R$ (ultimo 2: transacao + saldo)
        valores = RE_VALOR_RS.findall(linha)
        if len(valores) < 2:
            continue

        # Transacao = penultimo, saldo = ultimo
        valor_transacao = _parse_valor_br(valores[-2])
        # Determinar sinal do valor: checar texto antes do R$ da transacao
        # para ver se tem negativo

        # Extrair parte entre data e primeiro R$
        idx_primeiro_rs = linha.find('R$')
        if idx_primeiro_rs < 0:
            continue

        parte_texto = linha[len(data_raw):idx_primeiro_rs].strip()

        # Separar documento, historico e descricao
        # Formato: "DOCUMENTO HISTORICO DESCRICAO"
        documento, historico, descricao = _separar_doc_hist_desc_sisprime(parte_texto)

        # Determinar se e credito (entrada) ou debito (saida)
        is_credito = any(kw.lower() in historico.lower() for kw in CREDITO_KEYWORDS_SISPRIME)

        if is_credito:
            valor = abs(valor_transacao)
        else:
            valor = -abs(valor_transacao)

        lancamentos.append({
            "data": _parse_data_br(data_raw),
            "descricao": descricao if descricao else historico,
            "detalhes": historico if descricao else "",
            "documento": documento,
            "valor": valor,
            "historico": historico,
        })

    return lancamentos


def _separar_doc_hist_desc_sisprime(texto: str) -> tuple[str, str, str]:
    """Separa documento, historico e descricao do texto entre data e R$.

    Exemplos:
      'XXX018968 Débito Pix MARIO JOSE FERREIRA' -> ('XXX018968', 'Débito Pix', 'MARIO JOSE FERREIRA')
      '2025001031 Liq Parcela' -> ('2025001031', 'Liq Parcela', '')
      '37092326 Crédito Pix SOCIEDADE' -> ('37092326', 'Crédito Pix', 'SOCIEDADE')
      '5355 Fat cartão Sisprime 03/2026' -> ('5355', 'Fat cartão Sisprime', '03/2026')
      'Terreno Liq. Eletrônica IB ESTANCIA ALBATROZ' -> ('Terreno', 'Liq. Eletrônica IB', 'ESTANCIA ALBATROZ')
      'SALDO DEV Débito IOF' -> ('SALDO DEV', 'Débito IOF', '')
    """
    # Historico keywords para split
    # Longer/more specific patterns first to avoid partial matches
    HISTORICOS = [
        "Estorno Débito Pix", "Crédito Pix", "Pagamento Pix", "Débito Pix",
        "Fat cartão Sisprime", "Déb Cartão Sisprime", "Liq. Eletrônica IB",
        "Liq Parcela", "Débito DAS", "Débito IOF",
    ]

    for hist in HISTORICOS:
        idx = texto.find(hist)
        if idx >= 0:
            documento = texto[:idx].strip()
            resto = texto[idx + len(hist):].strip()
            return documento, hist, resto

    # Fallback: primeiro token = documento, resto = historico
    partes = texto.split(None, 1)
    if len(partes) == 2:
        return partes[0], partes[1], ""
    return texto, "", ""


KEYWORDS_ENCARGO_SISPRIME = ["IOF", "JUROS", "ENCARGO", "MULTA"]
KEYWORDS_TARIFA_SISPRIME = ["Déb Cartão Sisprime", "TARIFA"]


def _classificar_tipo_sisprime_extrato(historico: str, valor: float) -> str:
    """Retorna tipo semantico para lancamento de extrato Sisprime."""
    for kw in KEYWORDS_TARIFA_SISPRIME:
        if kw in historico:
            return "tarifa"

    hist_upper = historico.upper()
    for kw in KEYWORDS_ENCARGO_SISPRIME:
        if kw in hist_upper:
            return "encargo"

    return "entrada" if valor > 0 else "saida"


def _gerar_alertas_sisprime(lancamentos: list[dict], saldo_inicial: float) -> list[str]:
    """Gera alertas para extrato Sisprime."""
    alertas = []

    # Tentativas nao efetivadas
    nao_efetivados: dict[str, list[str]] = defaultdict(list)
    for lanc in lancamentos:
        if not lanc["efetivado"] and lanc["valor"] < 0:
            desc = lanc["descricao"]
            nao_efetivados[desc].append(lanc["data"])

    for desc, datas in nao_efetivados.items():
        for data in datas:
            data_fmt = _formatar_data_br(data)
            alertas.append(f"{desc}: estorno em {data_fmt}")

    # Saldo negativo
    saldo = saldo_inicial
    primeiro_negativo = None
    for lanc in lancamentos:
        if lanc["efetivado"]:
            saldo += lanc["valor"]
            if saldo < 0 and primeiro_negativo is None:
                primeiro_negativo = lanc["data"]

    if primeiro_negativo:
        data_fmt = _formatar_data_br(primeiro_negativo)
        alertas.append(f"Saldo ficou negativo em {data_fmt}")

    # IOF cobrado
    iof_total = sum(
        abs(l["valor"]) for l in lancamentos
        if "IOF" in l.get("detalhes", "").upper() or "IOF" in l.get("descricao", "").upper()
    )
    if iof_total > 0:
        alertas.append(f"IOF cobrado: R$ {iof_total:,.2f}")

    return alertas
